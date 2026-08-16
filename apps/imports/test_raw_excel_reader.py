from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase
from openpyxl import Workbook

from apps.imports.services.raw_excel_reader import (
    RawExcelReadError,
    read_raw_excel_rows,
)


class RawExcelReaderTests(SimpleTestCase):
    def make_upload(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append([None, None, None])
        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
            ]
        )
        worksheet.append(
            [
                "SOURCE A",
                10,
                "ARTICLE A",
            ]
        )
        worksheet.append([None, None, None])
        worksheet.append(
            [
                "SOURCE B",
                20,
                "ARTICLE B",
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        return SimpleUploadedFile(
            "chargement_raw.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    def test_rejects_duplicate_normalized_headers(self):
        workbook = Workbook()
        worksheet = workbook.active

        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                " qt\u00e9 ",
                "Article",
            ]
        )
        worksheet.append(
            [
                "SOURCE A",
                10,
                20,
                "ARTICLE A",
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "duplicate_headers.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            RawExcelReadError
        ) as context:
            read_raw_excel_rows(uploaded)

        self.assertEqual(
            context.exception.code,
            "duplicate_headers",
        )


    def test_rejects_empty_header_inside_columns(self):
        workbook = Workbook()
        worksheet = workbook.active

        worksheet.append(
            [
                "Vers l'emplacement",
                None,
                "Qt\u00e9",
                "Article",
            ]
        )
        worksheet.append(
            [
                "SOURCE A",
                "IGNORED",
                10,
                "ARTICLE A",
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "empty_header.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            RawExcelReadError
        ) as context:
            read_raw_excel_rows(uploaded)

        self.assertEqual(
            context.exception.code,
            "empty_headers",
        )
        self.assertEqual(
            context.exception.details["positions"],
            [2],
        )


    def test_rejects_multiple_worksheets(self):
        workbook = Workbook()

        first = workbook.active
        first.title = "Transferts"
        first.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
            ]
        )
        first.append(
            [
                "SOURCE A",
                10,
                "ARTICLE A",
            ]
        )

        second = workbook.create_sheet("Other")
        second.append(["Test"])
        second.append(["Value"])

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "multiple_sheets.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            RawExcelReadError
        ) as context:
            read_raw_excel_rows(uploaded)

        self.assertEqual(
            context.exception.code,
            "unexpected_worksheet_count",
        )
        self.assertEqual(
            context.exception.details["actual"],
            2,
        )


    def test_restores_uploaded_file_position_after_reading(self):
        uploaded = self.make_upload()
        uploaded.seek(7)

        read_raw_excel_rows(uploaded)

        self.assertEqual(
            uploaded.tell(),
            7,
        )


    def test_reads_raw_rows_and_preserves_excel_row_numbers(self):
        result = read_raw_excel_rows(
            self.make_upload(),
        )

        self.assertEqual(
            result.filename,
            "chargement_raw.xlsx",
        )
        self.assertEqual(
            result.worksheet_name,
            "Transferts",
        )
        self.assertEqual(
            result.header_row_number,
            2,
        )
        self.assertEqual(
            result.headers,
            (
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
            ),
        )

        self.assertEqual(
            len(result.rows),
            2,
        )

        self.assertEqual(
            result.rows[0].row_number,
            3,
        )
        self.assertEqual(
            result.rows[0].as_dict(),
            {
                "Vers l'emplacement": "SOURCE A",
                "Qt\u00e9": 10,
                "Article": "ARTICLE A",
            },
        )

        self.assertEqual(
            result.rows[1].row_number,
            5,
        )
        self.assertEqual(
            result.rows[1].as_dict(),
            {
                "Vers l'emplacement": "SOURCE B",
                "Qt\u00e9": 20,
                "Article": "ARTICLE B",
            },
        )
