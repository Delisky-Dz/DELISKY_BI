from io import BytesIO
from types import SimpleNamespace

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase
from openpyxl import Workbook

from apps.imports.services.report_row_cleaner import (
    STATUS_STOPPED,
    clean_report_rows,
    clean_report_rows_from_metadata,
)

from apps.imports.services.raw_chargement_file import (
    RawChargementFileError,
    adapt_raw_chargement_file,
    to_report_row_read_result,
)


class RawChargementFileTests(SimpleTestCase):
    def make_upload(
        self,
        *,
        second_source="SOURCE B",
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append([None, None, None])

        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
            ]
        )

        worksheet.append(
            [
                "SOURCE A",
                10,
                "ARTICLE A",
            ]
        )

        worksheet.append([None, None, None])

        worksheet.append(
            [
                second_source,
                20,
                "ARTICLE B",
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        return SimpleUploadedFile(
            "chargement_raw.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    def test_adapts_raw_excel_file_and_preserves_excel_row_numbers(self):
        result = adapt_raw_chargement_file(
            self.make_upload(),
            truck_mapping={
                "SOURCE A": "DELISKY LIV01",
                "SOURCE B": "NITA LIV01",
            },
        )

        self.assertEqual(
            result.filename,
            "chargement_raw.xlsx",
        )

        self.assertEqual(
            result.worksheet_name,
            "Transferts",
        )

        self.assertEqual(
            len(result.rows),
            2,
        )

        self.assertEqual(
            result.rows[0].excel_row_number,
            3,
        )

        self.assertEqual(
            result.rows[0].values,
            {
                "VAN": "DELISKY LIV01",
                "Qt\u00e9": 10,
                "Article": "ARTICLE A",
            },
        )

        self.assertEqual(
            result.rows[1].excel_row_number,
            5,
        )

        self.assertEqual(
            result.rows[1].values,
            {
                "VAN": "NITA LIV01",
                "Qt\u00e9": 20,
                "Article": "ARTICLE B",
            },
        )

    def test_preserves_possible_stopped_row_for_existing_cleaner(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
            ]
        )
        worksheet.append(
            [
                "SOURCE A",
                0,
                None,
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "chargement_stopped.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        result = adapt_raw_chargement_file(
            uploaded,
            truck_mapping={
                "SOURCE A": "DELISKY LIV01",
            },
        )

        self.assertEqual(
            len(result.rows),
            1,
        )

        self.assertEqual(
            result.rows[0].excel_row_number,
            2,
        )

        self.assertEqual(
            result.rows[0].values,
            {
                "VAN": "DELISKY LIV01",
                "Qt\u00e9": 0,
                "Article": None,
            },
        )


    def test_existing_cleaner_receives_real_excel_row_number(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
            ]
        )
        worksheet.append(
            [
                "SOURCE A",
                0,
                None,
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "chargement_stopped.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        adapted = adapt_raw_chargement_file(
            uploaded,
            truck_mapping={
                "SOURCE A": "DELISKY LIV01",
            },
        )

        row_result = to_report_row_read_result(
            adapted
        )

        preflight = SimpleNamespace(
            is_valid=True,
            parsed_filename=SimpleNamespace(
                report_type="CHARGEMENT",
            ),
        )

        cleaned = clean_report_rows(
            row_result,
            preflight,
        )

        self.assertEqual(
            cleaned.rows[0].row_number,
            2,
        )
        self.assertEqual(
            cleaned.rows[0].status,
            STATUS_STOPPED,
        )

        stopped_issue = next(
            issue
            for issue in cleaned.rows[0].issues
            if issue.code == "stopped_indicator"
        )

        self.assertFalse(
            stopped_issue.details["authoritative"]
        )


    def test_raw_chargement_cleans_without_period_metadata(self):
        adapted = adapt_raw_chargement_file(
            self.make_upload(),
            truck_mapping={
                "SOURCE A": "DELISKY LIV01",
                "SOURCE B": "NITA LIV01",
            },
        )

        row_result = to_report_row_read_result(
            adapted
        )

        cleaned = clean_report_rows_from_metadata(
            row_result
        )

        self.assertEqual(
            cleaned.report_type,
            "CHARGEMENT",
        )
        self.assertEqual(
            cleaned.total_rows,
            2,
        )
        self.assertEqual(
            cleaned.accepted_rows,
            2,
        )
        self.assertEqual(
            cleaned.rows[0].row_number,
            3,
        )
        self.assertEqual(
            cleaned.rows[1].row_number,
            5,
        )


    def test_report_row_result_preserves_raw_datetime_metadata(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
                "Date&Heure",
            ]
        )
        worksheet.append(
            [
                "SOURCE A",
                10,
                "ARTICLE A",
                "2026-08-05 10:30:00",
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "chargement_with_datetime.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        adapted = adapt_raw_chargement_file(
            uploaded,
            truck_mapping={
                "SOURCE A": "DELISKY LIV01",
            },
        )

        row_result = to_report_row_read_result(
            adapted
        )

        self.assertIn(
            "Date&Heure",
            row_result.headers,
        )
        self.assertEqual(
            row_result.rows[0].as_dict()[
                "Date&Heure"
            ],
            "2026-08-05 10:30:00",
        )

    def test_ignores_export_summary_footer_row(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Date&Heure",
                "Article",
                "Qt\u00e9",
                "De l'emplacement",
                "Vers l'emplacement",
                "Dernier achat",
                "D\u00e9tail",
                "Prix",
            ]
        )

        worksheet.append(
            [
                "2026-08-05 10:30:00",
                "ARTICLE A",
                10,
                "DEPOT",
                "SOURCE A",
                100,
                0,
                120,
            ]
        )

        worksheet.append(
            [
                None,
                "1",
                "10,000",
                None,
                None,
                "100,00",
                "0,00",
                "120,00",
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "chargement_with_footer.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        result = adapt_raw_chargement_file(
            uploaded,
            truck_mapping={
                "SOURCE A": "DELISKY LIV01",
            },
        )

        self.assertEqual(
            len(result.rows),
            1,
        )
        self.assertEqual(
            result.rows[0].excel_row_number,
            2,
        )

    def test_footer_like_row_before_end_is_not_ignored(
        self,
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Date&Heure",
                "Article",
                "Qt\u00e9",
                "De l'emplacement",
                "Vers l'emplacement",
            ]
        )

        worksheet.append(
            [
                "2026-08-05 10:30:00",
                "ARTICLE A",
                10,
                "DEPOT",
                "SOURCE A",
            ]
        )

        worksheet.append(
            [
                None,
                "1",
                "10,000",
                None,
                None,
            ]
        )

        worksheet.append(
            [
                "2026-08-05 11:00:00",
                "ARTICLE B",
                5,
                "DEPOT",
                "SOURCE A",
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "chargement_footer_like_middle.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            RawChargementFileError
        ) as context:
            adapt_raw_chargement_file(
                uploaded,
                truck_mapping={
                    "SOURCE A": "DELISKY LIV01",
                },
            )

        self.assertEqual(
            context.exception.details[
                "excel_row_number"
            ],
            3,
        )
        self.assertEqual(
            context.exception.details[
                "cause_code"
            ],
            "missing_source_truck_code",
        )

    def test_footer_with_wrong_row_count_is_not_ignored(
        self,
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Date&Heure",
                "Article",
                "Qt\u00e9",
                "De l'emplacement",
                "Vers l'emplacement",
            ]
        )

        worksheet.append(
            [
                "2026-08-05 10:30:00",
                "ARTICLE A",
                10,
                "DEPOT",
                "SOURCE A",
            ]
        )

        worksheet.append(
            [
                None,
                "2",
                "10,000",
                None,
                None,
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "chargement_footer_wrong_count.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            RawChargementFileError
        ) as context:
            adapt_raw_chargement_file(
                uploaded,
                truck_mapping={
                    "SOURCE A": "DELISKY LIV01",
                },
            )

        self.assertEqual(
            context.exception.details[
                "excel_row_number"
            ],
            3,
        )
        self.assertEqual(
            context.exception.details[
                "cause_code"
            ],
            "missing_source_truck_code",
        )

    def test_missing_truck_on_real_data_row_is_not_treated_as_footer(
        self,
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Date&Heure",
                "Article",
                "Qt\u00e9",
                "De l'emplacement",
                "Vers l'emplacement",
            ]
        )

        worksheet.append(
            [
                "2026-08-05 10:30:00",
                "ARTICLE A",
                10,
                "DEPOT",
                None,
            ]
        )

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "chargement_missing_truck.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            RawChargementFileError
        ) as context:
            adapt_raw_chargement_file(
                uploaded,
                truck_mapping={
                    "SOURCE A": "DELISKY LIV01",
                },
            )

        self.assertEqual(
            context.exception.code,
            "row_adaptation_failed",
        )
        self.assertEqual(
            context.exception.details[
                "excel_row_number"
            ],
            2,
        )
        self.assertEqual(
            context.exception.details[
                "cause_code"
            ],
            "missing_source_truck_code",
        )

    def test_error_reports_real_excel_row_number(self):
        with self.assertRaises(
            RawChargementFileError
        ) as context:
            adapt_raw_chargement_file(
                self.make_upload(
                    second_source="UNKNOWN VAN",
                ),
                truck_mapping={
                    "SOURCE A": "DELISKY LIV01",
                },
            )

        self.assertEqual(
            context.exception.code,
            "row_adaptation_failed",
        )

        self.assertEqual(
            context.exception.details[
                "excel_row_number"
            ],
            5,
        )

        self.assertEqual(
            context.exception.details[
                "cause_code"
            ],
            "source_truck_not_mapped",
        )
