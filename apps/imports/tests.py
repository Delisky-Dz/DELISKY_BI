from tempfile import TemporaryDirectory
from decimal import Decimal
from datetime import date, datetime
from datetime import date
from io import BytesIO, StringIO

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.db import IntegrityError, models, transaction
from django.test import SimpleTestCase, TestCase

from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

from apps.imports.models import (
    DistributionBrand as ReviewDistributionBrand,
    ImportBatch as ReviewImportBatch,
)

from apps.imports.models import (
    DistributionBrand as RowDistributionBrand,
    ImportBatch as RowImportBatch,
    ImportRow,
    ImportRowStatus,
)

from apps.imports.services import (
    create_or_update_import_review,
    ImportBatchReviewError,
    build_import_review_summary,
    build_import_review_summary_from_metadata,
    REVIEW_STATUS_REVIEWED,
    REVIEW_STATUS_BLOCKED,
    clean_report_rows,
    STATUS_STOPPED,
    STATUS_EXCLUDED,
    STATUS_ACCEPTED,
    parse_decimal_value,
    parse_datetime_value,
    parse_date_value,
    normalize_text,
    normalize_lookup_text,
    is_blank_value,
    ValueNormalizationError,
    ReportRowReadError,
    read_report_rows,
    run_import_preflight,
    ImportFilenameError,
    parse_import_filename,
    ExcelInspectionError,
    inspect_excel_file,
    validate_workbook_schema,
)

from .models import (
    DistributionBrand,
    ImportBatch,
    ImportBatchStatus,
    ImportReportType,
)


class DistributionBrandModelTests(TestCase):
    def test_save_normalizes_code_and_name(self):
        brand = DistributionBrand.objects.create(
            code="  bifa  ",
            name="  BIFA   Distribution  ",
        )

        self.assertEqual(brand.code, "BIFA")
        self.assertEqual(brand.name, "BIFA Distribution")

    def test_string_representation(self):
        brand = DistributionBrand(
            code="NITA",
            name="NITA",
        )

        self.assertEqual(str(brand), "NITA \u2014 NITA")

    def test_invalid_code_is_rejected(self):
        brand = DistributionBrand(
            code="NITA TEST",
            name="NITA Test",
        )

        with self.assertRaises(ValidationError):
            brand.full_clean()

    def test_blank_name_after_cleaning_is_rejected(self):
        brand = DistributionBrand(
            code="DELISKY",
            name="   ",
        )

        with self.assertRaises(ValidationError) as context:
            brand.full_clean()

        self.assertIn("name", context.exception.message_dict)

    def test_code_is_unique_case_insensitively(self):
        DistributionBrand.objects.create(
            code="BIFA",
            name="BIFA",
        )

        duplicate = DistributionBrand(
            code="bifa",
            name="Another BIFA",
        )

        with self.assertRaises(ValidationError):
            duplicate.full_clean()



class SeedBrandsCommandTests(TestCase):
    def run_seed_brands(self):
        output = StringIO()
        call_command(
            "seed_brands",
            stdout=output,
        )
        return output.getvalue()

    def test_command_creates_official_brands(self):
        self.run_seed_brands()

        self.assertEqual(
            set(
                DistributionBrand.objects.values_list(
                    "code",
                    flat=True,
                )
            ),
            {"BIFA", "DELISKY", "NITA"},
        )

    def test_command_is_idempotent(self):
        self.run_seed_brands()
        self.run_seed_brands()

        self.assertEqual(
            DistributionBrand.objects.count(),
            3,
        )

    def test_command_reactivates_disabled_official_brand(self):
        brand = DistributionBrand.objects.create(
            code="BIFA",
            name="BIFA",
            is_active=False,
        )

        self.run_seed_brands()

        brand.refresh_from_db()
        self.assertTrue(brand.is_active)

    def test_command_preserves_existing_official_brand_name(self):
        brand = DistributionBrand.objects.create(
            code="NITA",
            name="NITA Distribution",
            is_active=True,
        )

        self.run_seed_brands()

        brand.refresh_from_db()
        self.assertEqual(
            brand.name,
            "NITA Distribution",
        )

    def test_command_does_not_modify_custom_brands(self):
        custom_brand = DistributionBrand.objects.create(
            code="NOVA",
            name="NOVA",
            is_active=False,
        )

        self.run_seed_brands()

        custom_brand.refresh_from_db()
        self.assertFalse(custom_brand.is_active)
        self.assertEqual(custom_brand.name, "NOVA")



class ImportBatchModelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="import_batch_test_user",
            password="Temporary-Test-Password-2026",
        )
        cls.brand = DistributionBrand.objects.create(
            code="BIFA",
            name="BIFA",
        )
        cls.other_brand = DistributionBrand.objects.create(
            code="NITA",
            name="NITA",
        )

    def build_batch(self, **overrides):
        data = {
            "brand": self.brand,
            "report_type": ImportReportType.SALES,
            "period_start": date(2026, 3, 7),
            "period_end": date(2026, 3, 11),
            "original_filename": (
                "Sales_BIFA_2026-03-07_2026-03-11.xlsx"
            ),
            "file_size_bytes": 1024,
            "file_sha256": "a" * 64,
            "content_sha256": "b" * 64,
            "uploaded_by": self.user,
        }
        data.update(overrides)
        return ImportBatch(**data)

    def test_period_end_cannot_be_before_start(self):
        batch = self.build_batch(
            period_start=date(2026, 3, 11),
            period_end=date(2026, 3, 7),
        )

        with self.assertRaises(ValidationError) as context:
            batch.full_clean()

        self.assertIn(
            "period_end",
            context.exception.message_dict,
        )

    def test_opening_stock_requires_single_date(self):
        batch = self.build_batch(
            report_type=ImportReportType.OPENING_STOCK,
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 2),
        )

        with self.assertRaises(ValidationError) as context:
            batch.full_clean()

        self.assertIn(
            "period_end",
            context.exception.message_dict,
        )

    def test_row_counts_cannot_exceed_total(self):
        batch = self.build_batch(
            total_rows=10,
            accepted_rows=8,
            excluded_rows=3,
        )

        with self.assertRaises(ValidationError) as context:
            batch.full_clean()

        self.assertIn(
            "total_rows",
            context.exception.message_dict,
        )

    def test_approved_batch_cannot_have_blocking_errors(self):
        batch = self.build_batch(
            status=ImportBatchStatus.APPROVED,
            error_count=1,
        )

        with self.assertRaises(ValidationError) as context:
            batch.full_clean()

        self.assertIn(
            "status",
            context.exception.message_dict,
        )

    def test_save_normalizes_names_and_hashes(self):
        batch = self.build_batch(
            original_filename="  Sales_BIFA.xlsx  ",
            worksheet_name="  sales bifa  ",
            file_sha256="A" * 64,
            content_sha256="B" * 64,
        )
        batch.save()

        self.assertEqual(
            batch.original_filename,
            "Sales_BIFA.xlsx",
        )
        self.assertEqual(
            batch.worksheet_name,
            "sales bifa",
        )
        self.assertEqual(
            batch.file_sha256,
            "a" * 64,
        )
        self.assertEqual(
            batch.content_sha256,
            "b" * 64,
        )

    def test_pending_duplicate_file_hashes_are_allowed(self):
        first = self.build_batch()
        first.save()

        second = self.build_batch(
            content_sha256="c" * 64,
        )
        second.save()

        self.assertEqual(
            ImportBatch.objects.filter(
                file_sha256="a" * 64,
            ).count(),
            2,
        )

    def test_approved_file_hash_is_unique(self):
        self.build_batch(
            status=ImportBatchStatus.APPROVED,
        ).save()

        duplicate = self.build_batch(
            status=ImportBatchStatus.APPROVED,
            content_sha256="c" * 64,
        )

        with self.assertRaises(ValidationError):
            duplicate.full_clean()

    def test_approved_clean_content_is_unique_for_same_period(self):
        self.build_batch(
            status=ImportBatchStatus.APPROVED,
        ).save()

        duplicate = self.build_batch(
            status=ImportBatchStatus.APPROVED,
            file_sha256="c" * 64,
        )

        with self.assertRaises(ValidationError):
            duplicate.full_clean()

    def test_only_one_approved_opening_stock_per_month(self):
        self.build_batch(
            report_type=ImportReportType.OPENING_STOCK,
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 1),
            status=ImportBatchStatus.APPROVED,
        ).save()

        duplicate = self.build_batch(
            report_type=ImportReportType.OPENING_STOCK,
            period_start=date(2026, 3, 15),
            period_end=date(2026, 3, 15),
            status=ImportBatchStatus.APPROVED,
            file_sha256="c" * 64,
            content_sha256="d" * 64,
        )

        with self.assertRaises(ValidationError):
            duplicate.full_clean()

    def test_replacement_must_use_same_brand(self):
        original = self.build_batch(
            status=ImportBatchStatus.APPROVED,
        )
        original.save()

        replacement = self.build_batch(
            brand=self.other_brand,
            replaces_batch=original,
            file_sha256="c" * 64,
            content_sha256="d" * 64,
        )

        with self.assertRaises(ValidationError) as context:
            replacement.full_clean()

        self.assertIn(
            "replaces_batch",
            context.exception.message_dict,
        )

    def test_valid_replacement_for_same_identity_is_allowed(self):
        original = self.build_batch(
            status=ImportBatchStatus.APPROVED,
        )
        original.save()

        replacement = self.build_batch(
            replaces_batch=original,
            file_sha256="c" * 64,
            content_sha256="d" * 64,
        )

        replacement.full_clean()

        self.assertEqual(
            replacement.replaces_batch,
            original,
        )

    def test_string_representation_contains_identity(self):
        batch = self.build_batch()

        representation = str(batch)

        self.assertIn("BIFA", representation)
        self.assertIn("2026-03-07", representation)
        self.assertIn("2026-03-11", representation)



class ExcelInspectionServiceTests(SimpleTestCase):
    def make_excel_file(
        self,
        *,
        filename="Sales_TEST_2026-03-07_2026-03-11.xlsx",
        sheets=None,
    ):
        workbook = Workbook()
        default_sheet = workbook.active

        sheet_definitions = sheets or {
            "sales test": [
                ["VAN", "Date", "Client", "Montant"],
                ["TEST LIV01", "07/03/2026", "Client 1", 1000],
            ],
        }

        first_sheet = True

        for sheet_name, rows in sheet_definitions.items():
            if first_sheet:
                worksheet = default_sheet
                worksheet.title = sheet_name
                first_sheet = False
            else:
                worksheet = workbook.create_sheet(sheet_name)

            for row in rows:
                worksheet.append(row)

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        return SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

    def test_reads_valid_workbook_structure(self):
        uploaded = self.make_excel_file(
            sheets={
                "sales test": [
                    [None, None, None, None],
                    ["VAN", "Date", "Client", "Montant"],
                    ["TEST LIV01", "07/03/2026", "Client 1", 1000],
                    [None, None, None, None],
                    ["TEST LIV02", "08/03/2026", "Client 2", 500],
                ],
            },
        )

        result = inspect_excel_file(uploaded)
        worksheet = result.worksheets[0]

        self.assertEqual(result.worksheet_count, 1)
        self.assertEqual(worksheet.name, "sales test")
        self.assertEqual(worksheet.header_row_number, 2)
        self.assertEqual(
            worksheet.headers,
            ("VAN", "Date", "Client", "Montant"),
        )
        self.assertEqual(worksheet.column_count, 4)
        self.assertEqual(worksheet.data_row_count, 2)
        self.assertEqual(worksheet.blank_row_count, 1)

    def test_detects_empty_and_duplicate_headers(self):
        uploaded = self.make_excel_file(
            sheets={
                "items test": [
                    [
                        "VAN",
                        "",
                        "Client",
                        " client ",
                        None,
                    ],
                    [
                        "TEST LIV01",
                        "Article 1",
                        "Client 1",
                        "Client 1",
                        None,
                    ],
                ],
            },
        )

        result = inspect_excel_file(uploaded)
        worksheet = result.worksheets[0]

        self.assertEqual(
            worksheet.headers,
            ("VAN", "", "Client", "client"),
        )
        self.assertEqual(
            worksheet.empty_header_positions,
            (2,),
        )
        self.assertEqual(
            worksheet.duplicate_headers,
            ("client",),
        )

    def test_reads_multiple_worksheets(self):
        uploaded = self.make_excel_file(
            sheets={
                "sheet one": [
                    ["A", "B"],
                    [1, 2],
                ],
                "sheet two": [
                    ["C", "D"],
                    [3, 4],
                ],
            },
        )

        result = inspect_excel_file(uploaded)

        self.assertEqual(result.worksheet_count, 2)
        self.assertEqual(
            tuple(sheet.name for sheet in result.worksheets),
            ("sheet one", "sheet two"),
        )

    def test_rejects_unsupported_extension(self):
        uploaded = SimpleUploadedFile(
            "Sales_TEST.csv",
            b"VAN,Date,Client",
            content_type="text/csv",
        )

        with self.assertRaises(ExcelInspectionError) as context:
            inspect_excel_file(uploaded)

        self.assertEqual(
            context.exception.code,
            "unsupported_extension",
        )

    def test_rejects_empty_file(self):
        uploaded = SimpleUploadedFile(
            "Sales_TEST.xlsx",
            b"",
        )

        with self.assertRaises(ExcelInspectionError) as context:
            inspect_excel_file(uploaded)

        self.assertEqual(
            context.exception.code,
            "empty_file",
        )

    def test_rejects_invalid_xlsx_content(self):
        uploaded = SimpleUploadedFile(
            "Sales_TEST.xlsx",
            b"This is not a real Excel workbook.",
        )

        with self.assertRaises(ExcelInspectionError) as context:
            inspect_excel_file(uploaded)

        self.assertEqual(
            context.exception.code,
            "invalid_xlsx",
        )

    def test_rejects_file_over_size_limit(self):
        uploaded = self.make_excel_file()

        with self.assertRaises(ExcelInspectionError) as context:
            inspect_excel_file(
                uploaded,
                max_file_size_bytes=10,
            )

        self.assertEqual(
            context.exception.code,
            "file_too_large",
        )

    def test_restores_uploaded_file_position(self):
        uploaded = self.make_excel_file()
        uploaded.seek(7)

        inspect_excel_file(uploaded)

        self.assertEqual(uploaded.tell(), 7)



class ReportSchemaValidatorTests(SimpleTestCase):
    def make_inspection(
        self,
        headers,
        *,
        second_worksheet=False,
    ):
        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "report data"
        worksheet.append(headers)
        worksheet.append(
            [
                f"value-{position}"
                for position in range(1, len(headers) + 1)
            ]
        )

        if second_worksheet:
            second = workbook.create_sheet("other data")
            second.append(headers)
            second.append(
                [
                    f"other-{position}"
                    for position in range(
                        1,
                        len(headers) + 1,
                    )
                ]
            )

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        uploaded = SimpleUploadedFile(
            "Report_TEST.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

        return inspect_excel_file(uploaded)

    def test_accepts_valid_sales_headers(self):
        inspection = self.make_inspection(
            [
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ]
        )

        result = validate_workbook_schema(
            inspection,
            "SALES",
        )

        self.assertTrue(result.is_valid)
        self.assertEqual(result.errors, ())
        self.assertEqual(result.warnings, ())
        self.assertEqual(
            result.column_positions,
            (
                ("VAN", 1),
                ("Date&Heure", 2),
                ("Nom du client", 3),
                ("Total", 4),
                ("Region", 5),
            ),
        )

    def test_accepts_different_column_order(self):
        inspection = self.make_inspection(
            [
                "Total",
                "Region",
                "VAN",
                "Nom du client",
                "Date&Heure",
            ]
        )

        result = validate_workbook_schema(
            inspection,
            "SALES",
        )

        self.assertTrue(result.is_valid)
        self.assertEqual(
            result.column_positions,
            (
                ("VAN", 3),
                ("Date&Heure", 5),
                ("Nom du client", 4),
                ("Total", 1),
                ("Region", 2),
            ),
        )

    def test_rejects_missing_required_header(self):
        inspection = self.make_inspection(
            [
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
            ]
        )

        result = validate_workbook_schema(
            inspection,
            "SALES",
        )

        self.assertFalse(result.is_valid)
        self.assertIn(
            "missing_required_headers",
            tuple(issue.code for issue in result.errors),
        )
        self.assertEqual(
            result.errors[0].details["headers"],
            ["Region"],
        )

    def test_accepts_extra_header_with_warning(self):
        inspection = self.make_inspection(
            [
                "VAN",
                "Article",
                "Qt\u00e9",
                "Commercial",
            ]
        )

        result = validate_workbook_schema(
            inspection,
            "CHARGEMENT",
        )

        self.assertTrue(result.is_valid)
        self.assertEqual(result.errors, ())
        self.assertEqual(
            tuple(issue.code for issue in result.warnings),
            ("extra_headers",),
        )
        self.assertEqual(
            result.warnings[0].details["headers"],
            ["Commercial"],
        )

    def test_rejects_empty_header(self):
        inspection = self.make_inspection(
            [
                "VAN",
                "Date&Heure",
                "",
                "Total",
                "Region",
            ]
        )

        result = validate_workbook_schema(
            inspection,
            "SALES",
        )

        error_codes = tuple(
            issue.code
            for issue in result.errors
        )

        self.assertFalse(result.is_valid)
        self.assertIn("empty_headers", error_codes)
        self.assertIn(
            "missing_required_headers",
            error_codes,
        )

    def test_rejects_duplicate_header(self):
        inspection = self.make_inspection(
            [
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
                " region ",
            ]
        )

        result = validate_workbook_schema(
            inspection,
            "SALES",
        )

        self.assertFalse(result.is_valid)
        self.assertIn(
            "duplicate_headers",
            tuple(issue.code for issue in result.errors),
        )

    def test_rejects_multiple_worksheets(self):
        inspection = self.make_inspection(
            [
                "VAN",
                "Article",
                "Qt\u00e9 vendue",
                "Client",
            ],
            second_worksheet=True,
        )

        result = validate_workbook_schema(
            inspection,
            "ITEMS",
        )

        self.assertFalse(result.is_valid)
        self.assertEqual(
            tuple(issue.code for issue in result.errors),
            ("unexpected_worksheet_count",),
        )
        self.assertEqual(result.column_positions, ())
        self.assertEqual(result.worksheet_name, "")

    def test_normalizes_spaces_case_and_apostrophes(self):
        inspection = self.make_inspection(
            [
                " van ",
                "NOM   DU CLIENT",
                "Message d\u2019ignoration",
                " date ",
                "Cause d'ignoration",
            ]
        )

        result = validate_workbook_schema(
            inspection,
            "POS",
        )

        self.assertTrue(result.is_valid)
        self.assertEqual(result.errors, ())
        self.assertEqual(result.warnings, ())
        self.assertEqual(
            result.column_positions,
            (
                ("VAN", 1),
                ("Nom du client", 2),
                ("Message d'ignoration", 3),
                ("Date", 4),
                ("Cause d'ignoration", 5),
            ),
        )



class ImportFilenameParserTests(SimpleTestCase):
    def test_parses_period_report_filename(self):
        result = parse_import_filename(
            "Sales_BIFA_2026-03-07_2026-03-11.xlsx"
        )

        self.assertEqual(result.filename, (
            "Sales_BIFA_2026-03-07_2026-03-11.xlsx"
        ))
        self.assertEqual(result.report_type, "SALES")
        self.assertEqual(result.brand_code, "BIFA")
        self.assertEqual(
            result.period_start.isoformat(),
            "2026-03-07",
        )
        self.assertEqual(
            result.period_end.isoformat(),
            "2026-03-11",
        )

    def test_parses_opening_stock_with_one_date(self):
        result = parse_import_filename(
            "OpeningStock_DELISKY_2026-03-07.xlsx"
        )

        self.assertEqual(
            result.report_type,
            "OPENING_STOCK",
        )
        self.assertEqual(result.brand_code, "DELISKY")
        self.assertEqual(
            result.period_start,
            result.period_end,
        )

    def test_accepts_brand_with_underscore_and_hyphen(self):
        result = parse_import_filename(
            (
                "Items_BRAND_TEST-01_"
                "2026-03-07_2026-03-11.xlsx"
            )
        )

        self.assertEqual(
            result.brand_code,
            "BRAND_TEST-01",
        )
        self.assertEqual(result.report_type, "ITEMS")

    def test_uses_filename_only_from_full_path(self):
        result = parse_import_filename(
            (
                "C:/temporary/files/"
                "PoS_NITA_2026-03-07_2026-03-11.xlsx"
            )
        )

        self.assertEqual(
            result.filename,
            "PoS_NITA_2026-03-07_2026-03-11.xlsx",
        )
        self.assertEqual(result.report_type, "POS")
        self.assertEqual(result.brand_code, "NITA")

    def test_rejects_missing_period_end(self):
        with self.assertRaises(
            ImportFilenameError
        ) as context:
            parse_import_filename(
                "Sales_BIFA_2026-03-07.xlsx"
            )

        self.assertEqual(
            context.exception.code,
            "missing_period_end",
        )

    def test_rejects_opening_stock_with_two_dates(self):
        with self.assertRaises(
            ImportFilenameError
        ) as context:
            parse_import_filename(
                (
                    "OpeningStock_BIFA_"
                    "2026-03-01_2026-03-31.xlsx"
                )
            )

        self.assertEqual(
            context.exception.code,
            "unexpected_period_end",
        )

    def test_rejects_invalid_calendar_date(self):
        with self.assertRaises(
            ImportFilenameError
        ) as context:
            parse_import_filename(
                (
                    "Sales_BIFA_"
                    "2026-02-30_2026-03-11.xlsx"
                )
            )

        self.assertEqual(
            context.exception.code,
            "invalid_date",
        )
        self.assertEqual(
            context.exception.details["field"],
            "period_start",
        )

    def test_rejects_reversed_period(self):
        with self.assertRaises(
            ImportFilenameError
        ) as context:
            parse_import_filename(
                (
                    "Chargement_NITA_"
                    "2026-03-11_2026-03-07.xlsx"
                )
            )

        self.assertEqual(
            context.exception.code,
            "invalid_period",
        )

    def test_rejects_non_xlsx_filename(self):
        with self.assertRaises(
            ImportFilenameError
        ) as context:
            parse_import_filename(
                (
                    "Sales_BIFA_"
                    "2026-03-07_2026-03-11.csv"
                )
            )

        self.assertEqual(
            context.exception.code,
            "invalid_filename_format",
        )



class ImportPreflightTests(SimpleTestCase):
    def make_upload(
        self,
        filename,
        headers,
        *,
        second_worksheet=False,
    ):
        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "report data"
        worksheet.append(headers)
        worksheet.append(
            [
                f"value-{position}"
                for position in range(
                    1,
                    len(headers) + 1,
                )
            ]
        )

        if second_worksheet:
            second = workbook.create_sheet("other data")
            second.append(headers)
            second.append(
                [
                    f"other-{position}"
                    for position in range(
                        1,
                        len(headers) + 1,
                    )
                ]
            )

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        return SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

    def test_accepts_valid_sales_file(self):
        uploaded = self.make_upload(
            (
                "Sales_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            [
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
        )

        result = run_import_preflight(uploaded)

        self.assertTrue(result.is_valid)
        self.assertEqual(result.errors, ())
        self.assertEqual(result.warnings, ())
        self.assertEqual(
            result.parsed_filename.report_type,
            "SALES",
        )
        self.assertEqual(
            result.parsed_filename.brand_code,
            "BIFA",
        )
        self.assertEqual(
            result.inspection.worksheets[0].data_row_count,
            1,
        )
        self.assertEqual(
            dict(
                result.schema_validation.column_positions
            ),
            {
                "VAN": 1,
                "Date&Heure": 2,
                "Nom du client": 3,
                "Total": 4,
                "Region": 5,
            },
        )

    def test_accepts_valid_opening_stock_file(self):
        uploaded = self.make_upload(
            "OpeningStock_NITA_2026-03-07.xlsx",
            [
                "VAN",
                "Qt\u00e9",
                "Article",
            ],
        )

        result = run_import_preflight(uploaded)

        self.assertTrue(result.is_valid)
        self.assertEqual(
            result.parsed_filename.report_type,
            "OPENING_STOCK",
        )
        self.assertEqual(
            result.parsed_filename.period_start,
            result.parsed_filename.period_end,
        )

    def test_reports_invalid_filename(self):
        uploaded = self.make_upload(
            "Sales_BIFA_invalid.xlsx",
            [
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
        )

        result = run_import_preflight(uploaded)

        self.assertFalse(result.is_valid)
        self.assertIsNone(result.parsed_filename)
        self.assertIsNotNone(result.inspection)
        self.assertIsNone(result.schema_validation)
        self.assertEqual(
            tuple(
                (issue.stage, issue.code)
                for issue in result.errors
            ),
            (
                (
                    "filename",
                    "invalid_filename_format",
                ),
            ),
        )

    def test_reports_invalid_workbook_content(self):
        uploaded = SimpleUploadedFile(
            (
                "Sales_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            b"not a real Excel workbook",
        )

        result = run_import_preflight(uploaded)

        self.assertFalse(result.is_valid)
        self.assertIsNotNone(result.parsed_filename)
        self.assertIsNone(result.inspection)
        self.assertIsNone(result.schema_validation)
        self.assertEqual(
            tuple(
                (issue.stage, issue.code)
                for issue in result.errors
            ),
            (
                ("workbook", "invalid_xlsx"),
            ),
        )

    def test_reports_missing_required_column(self):
        uploaded = self.make_upload(
            (
                "Sales_DELISKY_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            [
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
            ],
        )

        result = run_import_preflight(uploaded)

        self.assertFalse(result.is_valid)
        self.assertEqual(
            tuple(
                (issue.stage, issue.code)
                for issue in result.errors
            ),
            (
                (
                    "schema",
                    "missing_required_headers",
                ),
            ),
        )
        self.assertEqual(
            result.errors[0].details["headers"],
            ["Region"],
        )

    def test_accepts_extra_column_with_warning(self):
        uploaded = self.make_upload(
            (
                "Chargement_BIFA_"
                "2026-03-07_2026-03-10.xlsx"
            ),
            [
                "VAN",
                "Qt\u00e9",
                "Article",
                "Commercial",
            ],
        )

        result = run_import_preflight(uploaded)

        self.assertTrue(result.is_valid)
        self.assertEqual(result.errors, ())
        self.assertEqual(
            tuple(
                (issue.stage, issue.code)
                for issue in result.warnings
            ),
            (
                ("schema", "extra_headers"),
            ),
        )
        self.assertEqual(
            result.warnings[0].details["headers"],
            ["Commercial"],
        )

    def test_rejects_multiple_worksheets(self):
        uploaded = self.make_upload(
            (
                "Items_NITA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            [
                "VAN",
                "Article",
                "Qt\u00e9 vendue",
                "Client",
            ],
            second_worksheet=True,
        )

        result = run_import_preflight(uploaded)

        self.assertFalse(result.is_valid)
        self.assertEqual(
            tuple(
                (issue.stage, issue.code)
                for issue in result.errors
            ),
            (
                (
                    "schema",
                    "unexpected_worksheet_count",
                ),
            ),
        )

    def test_accepts_reordered_columns(self):
        uploaded = self.make_upload(
            (
                "PoS_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            [
                "Date",
                "VAN",
                "Cause d'ignoration",
                "Nom du client",
                "Message d\u2019ignoration",
            ],
        )

        result = run_import_preflight(uploaded)

        self.assertTrue(result.is_valid)
        self.assertEqual(result.errors, ())
        self.assertEqual(
            dict(
                result.schema_validation.column_positions
            ),
            {
                "VAN": 2,
                "Nom du client": 4,
                "Message d'ignoration": 5,
                "Date": 1,
                "Cause d'ignoration": 3,
            },
        )



class ReportRowReaderTests(SimpleTestCase):
    def make_upload(
        self,
        *,
        filename=(
            "Sales_BIFA_"
            "2026-03-07_2026-03-11.xlsx"
        ),
        headers=None,
        rows=None,
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "report data"

        headers = headers or [
            "VAN",
            "Date&Heure",
            "Nom du client",
            "Total",
            "Region",
        ]

        rows = rows or [
            [
                "BIFA LIV01",
                "07/03/2026 09:10:11",
                "Client 1",
                1000.5,
                "MILA",
            ],
        ]

        worksheet.append(headers)

        for row in rows:
            worksheet.append(row)

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        return SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

    def test_reads_valid_rows(self):
        uploaded = self.make_upload(
            rows=[
                [
                    "BIFA LIV01",
                    "07/03/2026 09:10:11",
                    "Client 1",
                    1000.5,
                    "MILA",
                ],
                [
                    "BIFA LIV02",
                    "08/03/2026 10:20:30",
                    "Client 2",
                    2500,
                    "CHELGHOUM",
                ],
            ],
        )

        preflight = run_import_preflight(uploaded)
        result = read_report_rows(
            uploaded,
            preflight,
        )

        self.assertEqual(result.row_count, 2)
        self.assertEqual(result.report_type, "SALES")
        self.assertEqual(result.worksheet_name, "report data")
        self.assertEqual(
            result.headers,
            (
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ),
        )
        self.assertEqual(
            result.rows[0].row_number,
            2,
        )
        self.assertEqual(
            result.rows[1].row_number,
            3,
        )
        self.assertEqual(
            result.rows[0].as_dict()["Total"],
            1000.5,
        )

    def test_skips_blank_rows_and_keeps_excel_numbers(self):
        uploaded = self.make_upload(
            rows=[
                [
                    "BIFA LIV01",
                    "07/03/2026 09:10:11",
                    "Client 1",
                    1000,
                    "MILA",
                ],
                [None, None, None, None, None],
                [
                    "BIFA LIV02",
                    "08/03/2026 10:20:30",
                    "Client 2",
                    2000,
                    "MILA",
                ],
            ],
        )

        preflight = run_import_preflight(uploaded)
        result = read_report_rows(
            uploaded,
            preflight,
        )

        self.assertEqual(result.row_count, 2)
        self.assertEqual(
            tuple(
                row.row_number
                for row in result.rows
            ),
            (2, 4),
        )

    def test_maps_reordered_columns(self):
        uploaded = self.make_upload(
            headers=[
                "Total",
                "Region",
                "VAN",
                "Nom du client",
                "Date&Heure",
            ],
            rows=[
                [
                    1500,
                    "MILA",
                    "BIFA LIV01",
                    "Client 1",
                    "07/03/2026 09:10:11",
                ],
            ],
        )

        preflight = run_import_preflight(uploaded)
        result = read_report_rows(
            uploaded,
            preflight,
        )

        row = result.rows[0].as_dict()

        self.assertEqual(row["VAN"], "BIFA LIV01")
        self.assertEqual(
            row["Date&Heure"],
            "07/03/2026 09:10:11",
        )
        self.assertEqual(
            row["Nom du client"],
            "Client 1",
        )
        self.assertEqual(row["Total"], 1500)
        self.assertEqual(row["Region"], "MILA")

    def test_preserves_raw_values(self):
        uploaded = self.make_upload(
            rows=[
                [
                    "  BIFA LIV01  ",
                    "07/03/2026 09:10:11",
                    "  Client 1  ",
                    -25.5,
                    "",
                ],
            ],
        )

        preflight = run_import_preflight(uploaded)
        result = read_report_rows(
            uploaded,
            preflight,
        )

        row = result.rows[0].as_dict()

        self.assertEqual(
            row["VAN"],
            "  BIFA LIV01  ",
        )
        self.assertEqual(
            row["Nom du client"],
            "  Client 1  ",
        )
        self.assertEqual(row["Total"], -25.5)
        self.assertEqual(row["Region"], None)

    def test_restores_uploaded_file_position(self):
        uploaded = self.make_upload()
        uploaded.seek(7)

        preflight = run_import_preflight(uploaded)

        self.assertEqual(uploaded.tell(), 7)

        read_report_rows(
            uploaded,
            preflight,
        )

        self.assertEqual(uploaded.tell(), 7)

    def test_rejects_invalid_preflight(self):
        uploaded = self.make_upload(
            filename="Sales_BIFA_invalid.xlsx",
        )

        preflight = run_import_preflight(uploaded)

        self.assertFalse(preflight.is_valid)

        with self.assertRaises(
            ReportRowReadError
        ) as context:
            read_report_rows(
                uploaded,
                preflight,
            )

        self.assertEqual(
            context.exception.code,
            "invalid_preflight",
        )

    def test_supports_items_report(self):
        uploaded = self.make_upload(
            filename=(
                "Items_NITA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Article",
                "Qt\u00e9 vendue",
                "Client",
            ],
            rows=[
                [
                    "NITA LIV01",
                    "Article 1",
                    12,
                    "Client 1",
                ],
            ],
        )

        preflight = run_import_preflight(uploaded)
        result = read_report_rows(
            uploaded,
            preflight,
        )

        self.assertEqual(result.report_type, "ITEMS")
        self.assertEqual(result.row_count, 1)
        self.assertEqual(
            result.rows[0].as_dict(),
            {
                "VAN": "NITA LIV01",
                "Article": "Article 1",
                "Qt\u00e9 vendue": 12,
                "Client": "Client 1",
            },
        )



class ValueNormalizerTests(SimpleTestCase):
    def test_normalizes_text_and_blank_values(self):
        self.assertIsNone(normalize_text(None))
        self.assertIsNone(normalize_text("   "))

        self.assertEqual(
            normalize_text(
                "  Client\u00a0  d\u2019Alger  "
            ),
            "Client d'Alger",
        )

        self.assertTrue(is_blank_value(None))
        self.assertTrue(is_blank_value("   "))
        self.assertFalse(is_blank_value(0))

    def test_normalizes_lookup_text(self):
        self.assertEqual(
            normalize_lookup_text("  VAN Test  "),
            "van test",
        )
        self.assertIsNone(
            normalize_lookup_text("   ")
        )

    def test_parses_decimal_values(self):
        self.assertEqual(
            parse_decimal_value(1250),
            Decimal("1250"),
        )
        self.assertEqual(
            parse_decimal_value(1250.5),
            Decimal("1250.5"),
        )
        self.assertEqual(
            parse_decimal_value("1250,50"),
            Decimal("1250.50"),
        )
        self.assertEqual(
            parse_decimal_value(" -25.75 "),
            Decimal("-25.75"),
        )
        self.assertIsNone(
            parse_decimal_value(None)
        )

    def test_rejects_invalid_decimal_values(self):
        invalid_values = (
            True,
            "12A",
            "1,250.50",
            float("inf"),
        )

        for value in invalid_values:
            with self.subTest(value=value):
                with self.assertRaises(
                    ValueNormalizationError
                ) as context:
                    parse_decimal_value(value)

                self.assertEqual(
                    context.exception.code,
                    "invalid_number",
                )

    def test_parses_supported_datetime_strings(self):
        self.assertEqual(
            parse_datetime_value(
                "07/03/2026 09:10:11"
            ),
            datetime(2026, 3, 7, 9, 10, 11),
        )

        self.assertEqual(
            parse_datetime_value(
                "2026-03-07 09:10"
            ),
            datetime(2026, 3, 7, 9, 10),
        )

    def test_parses_date_only_values(self):
        self.assertEqual(
            parse_date_value("07/03/2026"),
            date(2026, 3, 7),
        )

        self.assertEqual(
            parse_date_value("2026-03-07"),
            date(2026, 3, 7),
        )

    def test_preserves_python_date_and_datetime_values(self):
        datetime_value = datetime(
            2026,
            3,
            7,
            9,
            10,
            11,
        )

        date_value = date(2026, 3, 7)

        self.assertEqual(
            parse_datetime_value(datetime_value),
            datetime_value,
        )

        self.assertEqual(
            parse_datetime_value(date_value),
            datetime(2026, 3, 7, 0, 0),
        )

    def test_parses_excel_serial_datetime(self):
        expected = datetime(
            2026,
            3,
            7,
            9,
            10,
            11,
        )

        excel_serial = to_excel(expected)

        self.assertEqual(
            parse_datetime_value(excel_serial),
            expected,
        )

    def test_returns_none_for_blank_dates(self):
        self.assertIsNone(
            parse_datetime_value(None)
        )
        self.assertIsNone(
            parse_datetime_value("   ")
        )
        self.assertIsNone(
            parse_date_value(None)
        )

    def test_rejects_invalid_datetime_values(self):
        invalid_values = (
            True,
            "31/02/2026",
            "not-a-date",
        )

        for value in invalid_values:
            with self.subTest(value=value):
                with self.assertRaises(
                    ValueNormalizationError
                ) as context:
                    parse_datetime_value(value)

                self.assertEqual(
                    context.exception.code,
                    "invalid_datetime",
                )



class ReportRowCleanerTests(SimpleTestCase):
    def make_cleaning_result(
        self,
        *,
        filename,
        headers,
        rows,
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "report data"
        worksheet.append(headers)

        for row in rows:
            worksheet.append(row)

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        uploaded = SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

        preflight = run_import_preflight(uploaded)
        self.assertTrue(
            preflight.is_valid,
            preflight.errors,
        )

        row_result = read_report_rows(
            uploaded,
            preflight,
        )

        return clean_report_rows(
            row_result,
            preflight,
        )

    def test_accepts_and_normalizes_sales_row(self):
        result = self.make_cleaning_result(
            filename=(
                "Sales_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
            rows=[
                [
                    "  BIFA   LIV01 ",
                    "07/03/2026 09:10:11",
                    "  Client   Test ",
                    1250.5,
                    " MILA  01 ",
                ],
            ],
        )

        row = result.rows[0]
        cleaned = row.cleaned_dict()

        self.assertEqual(row.status, STATUS_ACCEPTED)
        self.assertEqual(cleaned["van"], "BIFA LIV01")
        self.assertEqual(
            cleaned["van_normalized"],
            "bifa liv01",
        )
        self.assertEqual(
            cleaned["sale_datetime"],
            datetime(2026, 3, 7, 9, 10, 11),
        )
        self.assertEqual(
            cleaned["client"],
            "Client Test",
        )
        self.assertEqual(
            cleaned["total"],
            Decimal("1250.5"),
        )
        self.assertEqual(
            cleaned["region"],
            "MILA 01",
        )
        self.assertEqual(result.accepted_rows, 1)
        self.assertEqual(result.excluded_rows, 0)
        self.assertEqual(result.stopped_rows, 0)

    def test_marks_zero_sales_marker_as_stopped(self):
        result = self.make_cleaning_result(
            filename=(
                "Sales_DELISKY_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
            rows=[
                [
                    "DELISKY LIV03",
                    None,
                    None,
                    0,
                    None,
                ],
            ],
        )

        row = result.rows[0]

        self.assertEqual(row.status, STATUS_STOPPED)
        self.assertEqual(
            row.cleaned_dict()["total"],
            Decimal("0"),
        )
        self.assertEqual(
            tuple(issue.code for issue in row.issues),
            ("truck_stopped_for_period",),
        )
        self.assertTrue(
            row.issues[0].details["authoritative"]
        )
        self.assertEqual(result.stopped_rows, 1)
        self.assertEqual(result.warning_count, 1)
        self.assertEqual(result.error_count, 0)

    def test_excludes_sale_date_outside_filename_period(self):
        result = self.make_cleaning_result(
            filename=(
                "Sales_DELISKY_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
            rows=[
                [
                    "DELISKY LIV01",
                    "06/03/2026 10:00:00",
                    "Client Test",
                    1000,
                    "MILA",
                ],
            ],
        )

        row = result.rows[0]

        self.assertEqual(row.status, STATUS_EXCLUDED)
        self.assertIn(
            "date_outside_period",
            tuple(issue.code for issue in row.issues),
        )
        self.assertEqual(result.excluded_rows, 1)
        self.assertEqual(result.error_count, 1)

    def test_accepts_and_normalizes_items_row(self):
        result = self.make_cleaning_result(
            filename=(
                "Items_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Article",
                "Qt\u00e9 vendue",
                "Client",
            ],
            rows=[
                [
                    " BIFA LIV01 ",
                    " Article   Test ",
                    12,
                    " Client   Test ",
                ],
            ],
        )

        row = result.rows[0]
        cleaned = row.cleaned_dict()

        self.assertEqual(row.status, STATUS_ACCEPTED)
        self.assertEqual(
            cleaned["article"],
            "Article Test",
        )
        self.assertEqual(
            cleaned["quantity_sold"],
            Decimal("12"),
        )
        self.assertEqual(
            cleaned["client"],
            "Client Test",
        )

    def test_excludes_negative_item_quantity_with_warning(self):
        result = self.make_cleaning_result(
            filename=(
                "Items_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Article",
                "Qt\u00e9 vendue",
                "Client",
            ],
            rows=[
                [
                    "BIFA LIV01",
                    "Article Test",
                    -2,
                    "Client Test",
                ],
            ],
        )

        row = result.rows[0]

        self.assertEqual(row.status, STATUS_EXCLUDED)
        self.assertEqual(
            tuple(issue.code for issue in row.issues),
            ("negative_quantity",),
        )
        self.assertEqual(
            row.issues[0].severity,
            "WARNING",
        )
        self.assertEqual(result.warning_count, 1)
        self.assertEqual(result.error_count, 0)

    def test_marks_empty_items_activity_as_stopped(self):
        result = self.make_cleaning_result(
            filename=(
                "Items_NITA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Article",
                "Qt\u00e9 vendue",
                "Client",
            ],
            rows=[
                [
                    "NITA LIV01",
                    None,
                    None,
                    None,
                ],
            ],
        )

        row = result.rows[0]

        self.assertEqual(row.status, STATUS_STOPPED)
        self.assertEqual(
            tuple(issue.code for issue in row.issues),
            ("stopped_indicator",),
        )
        self.assertFalse(
            row.issues[0].details["authoritative"]
        )

    def test_marks_zero_stock_marker_as_stopped(self):
        result = self.make_cleaning_result(
            filename="OpeningStock_BIFA_2026-03-07.xlsx",
            headers=[
                "VAN",
                "Qt\u00e9",
                "Article",
            ],
            rows=[
                [
                    "BIFA PSLIV02",
                    0,
                    None,
                ],
            ],
        )

        row = result.rows[0]

        self.assertEqual(row.status, STATUS_STOPPED)
        self.assertEqual(
            row.cleaned_dict()["quantity"],
            Decimal("0"),
        )
        self.assertEqual(
            tuple(issue.code for issue in row.issues),
            ("stopped_indicator",),
        )

    def test_invalid_stock_quantity_is_not_stopped(self):
        result = self.make_cleaning_result(
            filename="OpeningStock_BIFA_2026-03-07.xlsx",
            headers=[
                "VAN",
                "Qt\u00e9",
                "Article",
            ],
            rows=[
                [
                    "BIFA LIV01",
                    "not-a-number",
                    None,
                ],
            ],
        )

        row = result.rows[0]
        issue_codes = tuple(
            issue.code
            for issue in row.issues
        )

        self.assertEqual(row.status, STATUS_EXCLUDED)
        self.assertNotEqual(row.status, STATUS_STOPPED)
        self.assertIn("invalid_number", issue_codes)
        self.assertIn("missing_article", issue_codes)

    def test_accepts_pos_numeric_message_with_warning(self):
        result = self.make_cleaning_result(
            filename=(
                "PoS_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Nom du client",
                "Message d'ignoration",
                "Date",
                "Cause d'ignoration",
            ],
            rows=[
                [
                    "BIFA LIV03",
                    "Client Test",
                    36.12469482,
                    "07/03/2026",
                    None,
                ],
            ],
        )

        row = result.rows[0]
        cleaned = row.cleaned_dict()

        self.assertEqual(row.status, STATUS_ACCEPTED)
        self.assertIsNone(
            cleaned["ignoration_message"]
        )
        self.assertEqual(
            cleaned["visit_date"],
            date(2026, 3, 7),
        )
        self.assertEqual(
            tuple(issue.code for issue in row.issues),
            ("numeric_ignoration_message",),
        )
        self.assertEqual(
            row.raw_dict()["Message d'ignoration"],
            36.12469482,
        )
        self.assertEqual(result.warning_count, 1)

    def test_marks_empty_pos_activity_as_stopped(self):
        result = self.make_cleaning_result(
            filename=(
                "PoS_NITA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Nom du client",
                "Message d'ignoration",
                "Date",
                "Cause d'ignoration",
            ],
            rows=[
                [
                    "NITA LIV01",
                    None,
                    None,
                    None,
                    None,
                ],
            ],
        )

        row = result.rows[0]

        self.assertEqual(row.status, STATUS_STOPPED)
        self.assertEqual(
            tuple(issue.code for issue in row.issues),
            ("stopped_indicator",),
        )



class ImportReviewSummaryTests(SimpleTestCase):
    def make_summary(
        self,
        *,
        filename,
        headers,
        rows,
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "report data"
        worksheet.append(headers)

        for row in rows:
            worksheet.append(row)

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        uploaded = SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

        preflight = run_import_preflight(uploaded)
        self.assertTrue(
            preflight.is_valid,
            preflight.errors,
        )

        row_result = read_report_rows(
            uploaded,
            preflight,
        )

        cleaning = clean_report_rows(
            row_result,
            preflight,
        )

        summary = build_import_review_summary(
            preflight,
            row_result,
            cleaning,
        )

        return summary

    def test_metadata_builder_matches_existing_preflight_builder(self):
        filename = (
            "Chargement_BIFA_"
            "2026-03-07_2026-03-11.xlsx"
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "report data"
        worksheet.append(
            [
                "VAN",
                "Qt\u00e9",
                "Article",
            ]
        )
        worksheet.append(
            [
                "BIFA LIV01",
                10,
                "ARTICLE A",
            ]
        )

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        uploaded = SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

        preflight = run_import_preflight(uploaded)
        self.assertTrue(
            preflight.is_valid,
            preflight.errors,
        )

        row_result = read_report_rows(
            uploaded,
            preflight,
        )
        cleaning = clean_report_rows(
            row_result,
            preflight,
        )

        existing_summary = build_import_review_summary(
            preflight,
            row_result,
            cleaning,
        )

        parsed = preflight.parsed_filename
        self.assertIsNotNone(parsed)

        metadata_summary = (
            build_import_review_summary_from_metadata(
                brand_code=parsed.brand_code,
                period_start=parsed.period_start,
                period_end=parsed.period_end,
                row_result=row_result,
                cleaning_result=cleaning,
            )
        )

        self.assertEqual(
            metadata_summary.as_dict(),
            existing_summary.as_dict(),
        )


    def test_clean_file_is_reviewed_and_approvable(self):
        summary = self.make_summary(
            filename=(
                "Sales_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
            rows=[
                [
                    "BIFA LIV01",
                    "07/03/2026 09:10:11",
                    "Client Test",
                    1250,
                    "MILA",
                ],
            ],
        )

        self.assertTrue(summary.can_approve)
        self.assertEqual(
            summary.recommended_status,
            REVIEW_STATUS_REVIEWED,
        )
        self.assertEqual(summary.total_rows, 1)
        self.assertEqual(summary.accepted_rows, 1)
        self.assertEqual(summary.excluded_rows, 0)
        self.assertEqual(summary.stopped_rows, 0)
        self.assertEqual(summary.retained_rows, 1)
        self.assertEqual(summary.warning_count, 0)
        self.assertEqual(summary.error_count, 0)
        self.assertEqual(summary.blocking_row_count, 0)
        self.assertEqual(summary.issue_groups, ())

    def test_warning_only_file_remains_approvable(self):
        summary = self.make_summary(
            filename=(
                "Items_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Article",
                "Qt\u00e9 vendue",
                "Client",
            ],
            rows=[
                [
                    "BIFA LIV01",
                    "Article Test",
                    -1,
                    "Client Test",
                ],
            ],
        )

        self.assertTrue(summary.can_approve)
        self.assertEqual(
            summary.recommended_status,
            REVIEW_STATUS_REVIEWED,
        )
        self.assertEqual(summary.accepted_rows, 0)
        self.assertEqual(summary.excluded_rows, 1)
        self.assertEqual(summary.warning_count, 1)
        self.assertEqual(summary.error_count, 0)
        self.assertEqual(summary.blocking_row_count, 0)

        group = summary.issue_groups[0]

        self.assertEqual(group.code, "negative_quantity")
        self.assertEqual(group.severity, "WARNING")
        self.assertEqual(group.count, 1)
        self.assertEqual(group.row_numbers, (2,))

    def test_error_file_is_blocked(self):
        summary = self.make_summary(
            filename=(
                "Sales_DELISKY_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
            rows=[
                [
                    "DELISKY LIV01",
                    "06/03/2026 10:00:00",
                    "Client Test",
                    1000,
                    "MILA",
                ],
            ],
        )

        self.assertFalse(summary.can_approve)
        self.assertEqual(
            summary.recommended_status,
            REVIEW_STATUS_BLOCKED,
        )
        self.assertEqual(summary.accepted_rows, 0)
        self.assertEqual(summary.excluded_rows, 1)
        self.assertEqual(summary.error_count, 1)
        self.assertEqual(summary.blocking_row_count, 1)

        group = summary.issue_groups[0]

        self.assertEqual(
            group.code,
            "date_outside_period",
        )
        self.assertEqual(group.severity, "ERROR")
        self.assertEqual(group.count, 1)
        self.assertEqual(group.row_numbers, (2,))

    def test_stopped_row_is_retained_and_approvable(self):
        summary = self.make_summary(
            filename=(
                "Sales_NITA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
            rows=[
                [
                    "NITA LIV01",
                    None,
                    None,
                    0,
                    None,
                ],
            ],
        )

        self.assertTrue(summary.can_approve)
        self.assertEqual(
            summary.recommended_status,
            REVIEW_STATUS_REVIEWED,
        )
        self.assertEqual(summary.accepted_rows, 0)
        self.assertEqual(summary.excluded_rows, 0)
        self.assertEqual(summary.stopped_rows, 1)
        self.assertEqual(summary.retained_rows, 1)
        self.assertEqual(summary.warning_count, 1)
        self.assertEqual(summary.error_count, 0)

        group = summary.issue_groups[0]

        self.assertEqual(
            group.code,
            "truck_stopped_for_period",
        )
        self.assertEqual(group.row_numbers, (2,))

    def test_groups_repeated_blocking_issues(self):
        summary = self.make_summary(
            filename=(
                "PoS_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Nom du client",
                "Message d'ignoration",
                "Date",
                "Cause d'ignoration",
            ],
            rows=[
                [
                    "BIFA LIV01",
                    "Client 1",
                    None,
                    "06/03/2026",
                    None,
                ],
                [
                    "BIFA LIV01",
                    "Client 2",
                    None,
                    "06/03/2026",
                    None,
                ],
            ],
        )

        self.assertFalse(summary.can_approve)
        self.assertEqual(
            summary.recommended_status,
            REVIEW_STATUS_BLOCKED,
        )
        self.assertEqual(summary.total_rows, 2)
        self.assertEqual(summary.excluded_rows, 2)
        self.assertEqual(summary.error_count, 2)
        self.assertEqual(summary.blocking_row_count, 2)

        self.assertEqual(
            len(summary.issue_groups),
            1,
        )

        group = summary.issue_groups[0]

        self.assertEqual(
            group.code,
            "date_outside_period",
        )
        self.assertEqual(group.count, 2)
        self.assertEqual(
            group.row_numbers,
            (2, 3),
        )

    def test_summary_serializes_to_plain_dictionary(self):
        summary = self.make_summary(
            filename=(
                "Chargement_NITA_"
                "2026-03-07_2026-03-10.xlsx"
            ),
            headers=[
                "VAN",
                "Qt\u00e9",
                "Article",
            ],
            rows=[
                [
                    "NITA LIV02",
                    10,
                    "Article Test",
                ],
            ],
        )

        payload = summary.as_dict()

        self.assertEqual(
            payload["report_type"],
            "CHARGEMENT",
        )
        self.assertEqual(
            payload["brand_code"],
            "NITA",
        )
        self.assertEqual(
            payload["period_start"],
            "2026-03-07",
        )
        self.assertEqual(
            payload["period_end"],
            "2026-03-10",
        )
        self.assertEqual(
            payload["recommended_status"],
            "REVIEWED",
        )
        self.assertEqual(
            payload["issue_groups"],
            [],
        )



class ImportBatchReviewServiceTests(TestCase):
    def setUp(self):
        self.media_directory = TemporaryDirectory()
        self.addCleanup(self.media_directory.cleanup)

        self.media_settings = self.settings(
            MEDIA_ROOT=self.media_directory.name,
        )
        self.media_settings.enable()
        self.addCleanup(self.media_settings.disable)

        self.user = get_user_model().objects.create_user(
            username="import-reviewer",
            password="test-password",
        )

        self.brand = ReviewDistributionBrand.objects.create(
            code="BIFA",
            name="BIFA",
            is_active=True,
        )

    def make_upload(
        self,
        *,
        filename=(
            "Sales_BIFA_"
            "2026-03-07_2026-03-11.xlsx"
        ),
        headers=None,
        rows=None,
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "report data"

        headers = headers or [
            "VAN",
            "Date&Heure",
            "Nom du client",
            "Total",
            "Region",
        ]

        rows = rows or [
            [
                "BIFA LIV01",
                "07/03/2026 09:10:11",
                "Client Test",
                1250,
                "MILA",
            ],
        ]

        worksheet.append(headers)

        for row in rows:
            worksheet.append(row)

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        return SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

    def test_creates_reviewed_batch_and_temporary_file(self):
        uploaded = self.make_upload()

        result = create_or_update_import_review(
            uploaded,
            uploaded_by=self.user,
        )

        self.assertTrue(result.created)
        self.assertEqual(
            ReviewImportBatch.objects.count(),
            1,
        )

        batch = result.batch
        batch.refresh_from_db()

        self.assertEqual(batch.status, "REVIEWED")
        self.assertEqual(batch.brand, self.brand)
        self.assertEqual(batch.report_type, "SALES")
        self.assertEqual(
            batch.period_start.isoformat(),
            "2026-03-07",
        )
        self.assertEqual(
            batch.period_end.isoformat(),
            "2026-03-11",
        )
        self.assertEqual(batch.total_rows, 1)
        self.assertEqual(batch.accepted_rows, 1)
        self.assertEqual(batch.excluded_rows, 0)
        self.assertEqual(batch.warning_count, 0)
        self.assertEqual(batch.error_count, 0)
        self.assertEqual(batch.uploaded_by, self.user)
        self.assertEqual(batch.reviewed_by, self.user)
        self.assertIsNotNone(batch.reviewed_at)
        self.assertIsNone(batch.approved_at)
        self.assertEqual(len(batch.file_sha256), 64)
        self.assertTrue(batch.source_file.name)
        self.assertTrue(
            batch.source_file.storage.exists(
                batch.source_file.name
            )
        )
        self.assertEqual(
            batch.review_summary["recommended_status"],
            "REVIEWED",
        )
        self.assertTrue(
            batch.review_summary["can_approve"]
        )

    def test_creates_blocked_batch_for_row_errors(self):
        uploaded = self.make_upload(
            rows=[
                [
                    "BIFA LIV01",
                    "06/03/2026 10:00:00",
                    "Client Test",
                    1000,
                    "MILA",
                ],
            ],
        )

        result = create_or_update_import_review(
            uploaded,
            uploaded_by=self.user,
        )

        batch = result.batch
        batch.refresh_from_db()

        self.assertEqual(batch.status, "BLOCKED")
        self.assertEqual(batch.total_rows, 1)
        self.assertEqual(batch.accepted_rows, 0)
        self.assertEqual(batch.excluded_rows, 1)
        self.assertEqual(batch.warning_count, 0)
        self.assertEqual(batch.error_count, 1)
        self.assertFalse(
            batch.review_summary["can_approve"]
        )
        self.assertEqual(
            batch.review_summary[
                "blocking_row_count"
            ],
            1,
        )
        self.assertEqual(
            batch.review_summary[
                "issue_groups"
            ][0]["code"],
            "date_outside_period",
        )

    def test_rejects_unknown_brand_without_creating_batch(self):
        uploaded = self.make_upload(
            filename=(
                "Sales_UNKNOWN_"
                "2026-03-07_2026-03-11.xlsx"
            ),
        )

        with self.assertRaises(
            ImportBatchReviewError
        ) as context:
            create_or_update_import_review(
                uploaded,
                uploaded_by=self.user,
            )

        self.assertEqual(
            context.exception.code,
            "unknown_brand",
        )
        self.assertEqual(
            ReviewImportBatch.objects.count(),
            0,
        )

    def test_rejects_invalid_preflight_without_batch(self):
        uploaded = self.make_upload(
            filename="Sales_BIFA_invalid.xlsx",
        )

        with self.assertRaises(
            ImportBatchReviewError
        ) as context:
            create_or_update_import_review(
                uploaded,
                uploaded_by=self.user,
            )

        self.assertEqual(
            context.exception.code,
            "preflight_failed",
        )
        self.assertEqual(
            ReviewImportBatch.objects.count(),
            0,
        )

    def test_updates_mutable_batch_and_replaces_file(self):
        first_result = create_or_update_import_review(
            self.make_upload(),
            uploaded_by=self.user,
        )

        batch_id = first_result.batch.pk
        old_file_name = first_result.batch.source_file.name
        storage = first_result.batch.source_file.storage

        second_upload = self.make_upload(
            rows=[
                [
                    "BIFA LIV02",
                    "08/03/2026 11:00:00",
                    "Another Client",
                    2500,
                    "MILA",
                ],
                [
                    "BIFA LIV03",
                    "09/03/2026 12:00:00",
                    "Third Client",
                    3000,
                    "MILA",
                ],
            ],
        )

        with self.captureOnCommitCallbacks(
            execute=True
        ):
            second_result = (
                create_or_update_import_review(
                    second_upload,
                    uploaded_by=self.user,
                    reviewed_by=self.user,
                    batch=first_result.batch,
                )
            )

        self.assertFalse(second_result.created)
        self.assertEqual(
            second_result.batch.pk,
            batch_id,
        )

        second_result.batch.refresh_from_db()
        new_file_name = (
            second_result.batch.source_file.name
        )

        self.assertEqual(
            ReviewImportBatch.objects.count(),
            1,
        )
        self.assertEqual(
            second_result.batch.total_rows,
            2,
        )
        self.assertEqual(
            second_result.batch.accepted_rows,
            2,
        )
        self.assertNotEqual(
            old_file_name,
            new_file_name,
        )
        self.assertFalse(
            storage.exists(old_file_name)
        )
        self.assertTrue(
            storage.exists(new_file_name)
        )

    def test_rejects_update_of_approved_batch(self):
        approved = ReviewImportBatch.objects.create(
            brand=self.brand,
            report_type="SALES",
            period_start=date(2026, 3, 7),
            period_end=date(2026, 3, 11),
            original_filename=(
                "Sales_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            status="APPROVED",
            total_rows=1,
            accepted_rows=1,
            excluded_rows=0,
            warning_count=0,
            error_count=0,
            uploaded_by=self.user,
            reviewed_by=self.user,
            approved_by=self.user,
        )

        with self.assertRaises(
            ImportBatchReviewError
        ) as context:
            create_or_update_import_review(
                self.make_upload(),
                uploaded_by=self.user,
                batch=approved,
            )

        self.assertEqual(
            context.exception.code,
            "immutable_batch",
        )

        approved.refresh_from_db()
        self.assertEqual(
            approved.status,
            "APPROVED",
        )

    def test_rejects_unsaved_user(self):
        unsaved_user = get_user_model()(
            username="unsaved-reviewer"
        )

        with self.assertRaises(
            ImportBatchReviewError
        ) as context:
            create_or_update_import_review(
                self.make_upload(),
                uploaded_by=unsaved_user,
            )

        self.assertEqual(
            context.exception.code,
            "unsaved_user",
        )
        self.assertEqual(
            ReviewImportBatch.objects.count(),
            0,
        )


    def test_persists_stopped_row_count(self):
        uploaded = self.make_upload(
            rows=[
                [
                    "BIFA LIV09",
                    None,
                    None,
                    0,
                    None,
                ],
            ],
        )

        result = create_or_update_import_review(
            uploaded,
            uploaded_by=self.user,
        )

        result.batch.refresh_from_db()

        self.assertEqual(
            result.batch.status,
            "REVIEWED",
        )
        self.assertEqual(result.batch.total_rows, 1)
        self.assertEqual(result.batch.accepted_rows, 0)
        self.assertEqual(result.batch.excluded_rows, 0)
        self.assertEqual(result.batch.stopped_rows, 1)
        self.assertEqual(
            result.batch.review_summary["stopped_rows"],
            1,
        )



class ImportRowModelTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="row-model-user",
            password="test-password",
        )

        self.brand = RowDistributionBrand.objects.create(
            code="ROWTEST",
            name="Row Test",
            is_active=True,
        )

        self.batch = RowImportBatch.objects.create(
            brand=self.brand,
            report_type="SALES",
            period_start=date(2026, 3, 7),
            period_end=date(2026, 3, 11),
            original_filename=(
                "Sales_ROWTEST_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            file_sha256="a" * 64,
            status="PENDING",
            uploaded_by=self.user,
        )

    def make_row(
        self,
        *,
        excel_row_number=2,
        status=ImportRowStatus.ACCEPTED,
        row_sha256=None,
        raw_data=None,
        cleaned_data=None,
        issues=None,
    ):
        return ImportRow.objects.create(
            batch=self.batch,
            excel_row_number=excel_row_number,
            status=status,
            raw_data=(
                raw_data
                if raw_data is not None
                else {
                    "VAN": "ROWTEST LIV01",
                    "Total": 1250.5,
                }
            ),
            cleaned_data=(
                cleaned_data
                if cleaned_data is not None
                else {
                    "van": "ROWTEST LIV01",
                    "total": Decimal("1250.5"),
                }
            ),
            issues=(
                issues
                if issues is not None
                else []
            ),
            row_sha256=(
                row_sha256
                if row_sha256 is not None
                else "b" * 64
            ),
        )

    def test_creates_import_row(self):
        row = self.make_row()

        self.assertEqual(row.batch, self.batch)
        self.assertEqual(row.excel_row_number, 2)
        self.assertEqual(
            row.status,
            ImportRowStatus.ACCEPTED,
        )
        self.assertEqual(
            row.raw_data["VAN"],
            "ROWTEST LIV01",
        )
        self.assertEqual(
            row.cleaned_data["total"],
            Decimal("1250.5"),
        )
        self.assertEqual(row.issues, [])

    def test_json_encoder_supports_decimal_and_datetime(self):
        row = self.make_row(
            cleaned_data={
                "total": Decimal("1250.50"),
                "sale_datetime": datetime(
                    2026,
                    3,
                    7,
                    9,
                    10,
                    11,
                ),
            },
        )

        row.refresh_from_db()

        self.assertEqual(
            row.cleaned_data["total"],
            "1250.50",
        )
        self.assertEqual(
            row.cleaned_data["sale_datetime"],
            "2026-03-07T09:10:11",
        )

    def test_normalizes_row_hash_to_lowercase(self):
        row = self.make_row(
            row_sha256="ABCDEF" * 10 + "ABCD",
        )

        self.assertEqual(
            row.row_sha256,
            ("abcdef" * 10) + "abcd",
        )

    def test_rejects_duplicate_excel_row_in_same_batch(self):
        self.make_row(
            excel_row_number=2,
            row_sha256="c" * 64,
        )

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.make_row(
                    excel_row_number=2,
                    row_sha256="d" * 64,
                )

    def test_allows_same_excel_row_in_different_batch(self):
        self.make_row(
            excel_row_number=2,
            row_sha256="e" * 64,
        )

        second_batch = RowImportBatch.objects.create(
            brand=self.brand,
            report_type="SALES",
            period_start=date(2026, 3, 12),
            period_end=date(2026, 3, 13),
            original_filename=(
                "Sales_ROWTEST_"
                "2026-03-12_2026-03-13.xlsx"
            ),
            file_sha256="f" * 64,
            status="PENDING",
            uploaded_by=self.user,
        )

        second_row = ImportRow.objects.create(
            batch=second_batch,
            excel_row_number=2,
            status=ImportRowStatus.ACCEPTED,
            raw_data={"VAN": "ROWTEST LIV02"},
            cleaned_data={"van": "ROWTEST LIV02"},
            issues=[],
            row_sha256="1" * 64,
        )

        self.assertIsNotNone(second_row.pk)

    def test_database_rejects_excel_row_number_below_two(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.make_row(
                    excel_row_number=1,
                    row_sha256="2" * 64,
                )

    def test_clean_rejects_invalid_json_container_types(self):
        row = ImportRow(
            batch=self.batch,
            excel_row_number=2,
            status=ImportRowStatus.EXCLUDED,
            raw_data=[],
            cleaned_data=[],
            issues={},
            row_sha256="3" * 64,
        )

        with self.assertRaises(ValidationError) as context:
            row.full_clean()

        self.assertIn(
            "raw_data",
            context.exception.message_dict,
        )
        self.assertIn(
            "cleaned_data",
            context.exception.message_dict,
        )
        self.assertIn(
            "issues",
            context.exception.message_dict,
        )

    def test_deleting_batch_cascades_to_rows(self):
        row = self.make_row(
            row_sha256="4" * 64,
        )
        row_id = row.pk

        self.batch.delete()

        self.assertFalse(
            ImportRow.objects.filter(
                pk=row_id,
            ).exists()
        )



class ImportRowStagingPersistenceTests(TestCase):
    def setUp(self):
        self.media_directory = TemporaryDirectory()
        self.addCleanup(self.media_directory.cleanup)

        self.media_settings = self.settings(
            MEDIA_ROOT=self.media_directory.name,
        )
        self.media_settings.enable()
        self.addCleanup(self.media_settings.disable)

        self.user = get_user_model().objects.create_user(
            username="row-staging-user",
            password="test-password",
        )

        self.brand = ReviewDistributionBrand.objects.create(
            code="BIFA",
            name="BIFA",
            is_active=True,
        )

    def make_upload(
        self,
        *,
        filename,
        headers,
        rows,
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "report data"
        worksheet.append(headers)

        for row in rows:
            worksheet.append(row)

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        return SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

    def test_review_persists_row_and_content_hash(self):
        uploaded = self.make_upload(
            filename=(
                "Sales_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
            rows=[
                [
                    "BIFA LIV01",
                    "07/03/2026 09:10:11",
                    "Client Test",
                    1250.5,
                    "MILA",
                ],
            ],
        )

        result = create_or_update_import_review(
            uploaded,
            uploaded_by=self.user,
        )

        batch = result.batch
        batch.refresh_from_db()

        self.assertEqual(batch.rows.count(), 1)
        self.assertEqual(len(batch.content_sha256), 64)
        self.assertNotEqual(batch.content_sha256, "")

        row = batch.rows.get()

        self.assertEqual(row.excel_row_number, 2)
        self.assertEqual(
            row.status,
            ImportRowStatus.ACCEPTED,
        )
        self.assertEqual(len(row.row_sha256), 64)
        self.assertEqual(
            row.raw_data["VAN"],
            "BIFA LIV01",
        )
        self.assertEqual(
            row.cleaned_data["total"],
            "1250.5",
        )
        self.assertEqual(
            row.cleaned_data["sale_datetime"],
            "2026-03-07T09:10:11",
        )

    def test_persists_all_three_row_statuses(self):
        uploaded = self.make_upload(
            filename=(
                "Items_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Article",
                "Qt\u00e9 vendue",
                "Client",
            ],
            rows=[
                [
                    "BIFA LIV01",
                    "Article 1",
                    2,
                    "Client 1",
                ],
                [
                    "BIFA LIV01",
                    "Article 2",
                    -1,
                    "Client 2",
                ],
                [
                    "BIFA LIV09",
                    None,
                    None,
                    None,
                ],
            ],
        )

        result = create_or_update_import_review(
            uploaded,
            uploaded_by=self.user,
        )

        batch = result.batch
        batch.refresh_from_db()

        self.assertEqual(batch.total_rows, 3)
        self.assertEqual(batch.accepted_rows, 1)
        self.assertEqual(batch.excluded_rows, 1)
        self.assertEqual(batch.stopped_rows, 1)
        self.assertEqual(batch.rows.count(), 3)

        statuses = {
            status: count
            for status, count in (
                batch.rows
                .values_list("status")
                .order_by("status")
                .annotate(count=models.Count("id"))
            )
        }

        self.assertEqual(
            statuses,
            {
                ImportRowStatus.ACCEPTED: 1,
                ImportRowStatus.EXCLUDED: 1,
                ImportRowStatus.STOPPED: 1,
            },
        )

        excluded = batch.rows.get(
            status=ImportRowStatus.EXCLUDED,
        )

        self.assertEqual(
            excluded.issues[0]["code"],
            "negative_quantity",
        )
        self.assertEqual(
            excluded.cleaned_data["quantity_sold"],
            "-1",
        )

    def test_blocked_batch_still_persists_review_rows(self):
        uploaded = self.make_upload(
            filename=(
                "Sales_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
            rows=[
                [
                    "BIFA LIV01",
                    "06/03/2026 10:00:00",
                    "Client Outside",
                    1000,
                    "MILA",
                ],
                [
                    "BIFA LIV02",
                    "07/03/2026 11:00:00",
                    "Client Valid",
                    2000,
                    "MILA",
                ],
            ],
        )

        result = create_or_update_import_review(
            uploaded,
            uploaded_by=self.user,
        )

        batch = result.batch
        batch.refresh_from_db()

        self.assertEqual(batch.status, "BLOCKED")
        self.assertEqual(batch.rows.count(), 2)
        self.assertEqual(
            batch.rows.filter(
                status=ImportRowStatus.EXCLUDED,
            ).count(),
            1,
        )
        self.assertEqual(
            batch.rows.filter(
                status=ImportRowStatus.ACCEPTED,
            ).count(),
            1,
        )

        excluded = batch.rows.get(
            status=ImportRowStatus.EXCLUDED,
        )

        self.assertEqual(
            excluded.issues[0]["code"],
            "date_outside_period",
        )

    def test_updating_review_replaces_old_rows(self):
        first_upload = self.make_upload(
            filename=(
                "Sales_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
            rows=[
                [
                    "BIFA LIV01",
                    "07/03/2026 09:00:00",
                    "Client 1",
                    1000,
                    "MILA",
                ],
                [
                    "BIFA LIV02",
                    "08/03/2026 10:00:00",
                    "Client 2",
                    2000,
                    "MILA",
                ],
            ],
        )

        first_result = create_or_update_import_review(
            first_upload,
            uploaded_by=self.user,
        )

        batch = first_result.batch
        old_hash = batch.content_sha256

        self.assertEqual(batch.rows.count(), 2)

        second_upload = self.make_upload(
            filename=(
                "Sales_BIFA_"
                "2026-03-07_2026-03-11.xlsx"
            ),
            headers=[
                "VAN",
                "Date&Heure",
                "Nom du client",
                "Total",
                "Region",
            ],
            rows=[
                [
                    "BIFA LIV03",
                    "09/03/2026 12:00:00",
                    "Replacement Client",
                    3000,
                    "MILA",
                ],
            ],
        )

        second_result = create_or_update_import_review(
            second_upload,
            uploaded_by=self.user,
            batch=batch,
        )

        second_result.batch.refresh_from_db()

        self.assertFalse(second_result.created)
        self.assertEqual(
            second_result.batch.rows.count(),
            1,
        )
        self.assertEqual(
            ImportRow.objects.filter(
                batch=second_result.batch,
            ).count(),
            1,
        )
        self.assertNotEqual(
            second_result.batch.content_sha256,
            old_hash,
        )

        row = second_result.batch.rows.get()

        self.assertEqual(row.excel_row_number, 2)
        self.assertEqual(
            row.cleaned_data["client"],
            "Replacement Client",
        )
