from io import BytesIO
from tempfile import TemporaryDirectory

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook

from apps.fleet.models import Truck
from apps.imports.models import (
    DistributionBrand,
    ImportBatch,
    ImportSourceSystem,
    ImportSourceUpload,
    SourceTruckMapping,
)
from apps.imports.services.raw_chargement_derived_review import (
    create_raw_chargement_derived_import_reviews,
)
from apps.imports.services.raw_chargement_review import (
    RawChargementImportReviewError,
)


class RawChargementDerivedReviewTests(TestCase):
    def setUp(self):
        self.media_directory = TemporaryDirectory()
        self.addCleanup(self.media_directory.cleanup)

        self.media_settings = self.settings(
            MEDIA_ROOT=self.media_directory.name,
        )
        self.media_settings.enable()
        self.addCleanup(self.media_settings.disable)

        self.user = get_user_model().objects.create_user(
            username="raw-derived-reviewer",
            password="test-password",
        )

        self.delisky = DistributionBrand.objects.create(
            code="DELISKY",
            name="DELISKY",
            is_active=True,
        )

        self.nita = DistributionBrand.objects.create(
            code="NITA",
            name="NITA",
            is_active=True,
        )

        self.delisky_truck = Truck.objects.create(
            internal_code="DELISKY LIV01",
            distribution_brand=self.delisky,
            registration_number="DERIVED-DELISKY-01",
            brand="TEST BRAND",
            model="TEST MODEL",
        )

        self.nita_truck = Truck.objects.create(
            internal_code="NITA LIV01",
            distribution_brand=self.nita,
            registration_number="DERIVED-NITA-01",
            brand="TEST BRAND",
            model="TEST MODEL",
        )

        self.source_system = (
            ImportSourceSystem.objects.create(
                code="AIO_WEB",
                name="AIO-WEB",
                is_active=True,
            )
        )

        SourceTruckMapping.objects.create(
            source_system=self.source_system,
            source_code="VAN1-DELISKY",
            truck=self.delisky_truck,
            is_active=True,
        )

        SourceTruckMapping.objects.create(
            source_system=self.source_system,
            source_code="VAN1-NITA",
            truck=self.nita_truck,
            is_active=True,
        )

    def make_upload(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
            ]
        )

        worksheet.append(
            [
                "VAN1-DELISKY",
                10,
                "ARTICLE DELISKY",
            ]
        )

        worksheet.append(
            [
                "VAN1-NITA",
                20,
                "ARTICLE NITA",
            ]
        )

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        return SimpleUploadedFile(
            "aio_mixed_raw.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    def test_allows_blank_datetime_for_stopped_chargement_row(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
                "Date&Heure",
            ]
        )

        worksheet.append(
            [
                "VAN1-DELISKY",
                0,
                None,
                None,
            ]
        )

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        upload = SimpleUploadedFile(
            "stopped_blank_datetime.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        result = create_raw_chargement_derived_import_reviews(
            upload,
            source_system_code="AIO_WEB",
            uploaded_by=self.user,
            period_start="2026-08-01",
            period_end="2026-08-17",
            original_filename="stopped_blank_datetime.xlsx",
        )

        self.assertEqual(
            len(result.batches),
            1,
        )
        self.assertEqual(
            ImportSourceUpload.objects.count(),
            1,
        )
        self.assertEqual(
            ImportBatch.objects.count(),
            1,
        )

    def test_rejects_blank_datetime_on_active_chargement_row(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
                "Date&Heure",
            ]
        )

        worksheet.append(
            [
                "VAN1-DELISKY",
                10,
                "ARTICLE DELISKY",
                None,
            ]
        )

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        upload = SimpleUploadedFile(
            "blank_datetime.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            RawChargementImportReviewError
        ) as context:
            create_raw_chargement_derived_import_reviews(
                upload,
                source_system_code="AIO_WEB",
                uploaded_by=self.user,
                period_start="2026-08-01",
                period_end="2026-08-17",
                original_filename="blank_datetime.xlsx",
            )

        self.assertEqual(
            context.exception.code,
            "missing_datetime",
        )

        self.assertEqual(
            context.exception.details[
                "excel_row_number"
            ],
            2,
        )

        self.assertEqual(
            ImportSourceUpload.objects.count(),
            0,
        )
        self.assertEqual(
            ImportBatch.objects.count(),
            0,
        )

    def test_rejects_invalid_raw_datetime_before_persistence(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
                "Date&Heure",
            ]
        )

        worksheet.append(
            [
                "VAN1-DELISKY",
                10,
                "ARTICLE DELISKY",
                "NOT-A-DATETIME",
            ]
        )

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        upload = SimpleUploadedFile(
            "invalid_datetime.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            RawChargementImportReviewError
        ) as context:
            create_raw_chargement_derived_import_reviews(
                upload,
                source_system_code="AIO_WEB",
                uploaded_by=self.user,
                period_start="2026-08-01",
                period_end="2026-08-17",
                original_filename="invalid_datetime.xlsx",
            )

        self.assertEqual(
            context.exception.code,
            "invalid_datetime",
        )

        self.assertEqual(
            context.exception.details[
                "excel_row_number"
            ],
            2,
        )

        self.assertEqual(
            ImportSourceUpload.objects.count(),
            0,
        )
        self.assertEqual(
            ImportBatch.objects.count(),
            0,
        )

    def test_rejects_raw_datetime_outside_declared_period(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Transferts"

        worksheet.append(
            [
                "Vers l'emplacement",
                "Qt\u00e9",
                "Article",
                "Date&Heure",
            ]
        )

        worksheet.append(
            [
                "VAN1-DELISKY",
                10,
                "ARTICLE DELISKY",
                "2026-07-31 10:30:00",
            ]
        )

        output = BytesIO()
        workbook.save(output)
        workbook.close()

        upload = SimpleUploadedFile(
            "outside_period.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            RawChargementImportReviewError
        ) as context:
            create_raw_chargement_derived_import_reviews(
                upload,
                source_system_code="AIO_WEB",
                uploaded_by=self.user,
                period_start="2026-08-01",
                period_end="2026-08-17",
                original_filename="outside_period.xlsx",
            )

        self.assertEqual(
            context.exception.code,
            "date_outside_period",
        )

        self.assertEqual(
            ImportSourceUpload.objects.count(),
            0,
        )
        self.assertEqual(
            ImportBatch.objects.count(),
            0,
        )

    def test_one_mixed_raw_file_creates_source_upload_and_brand_batches(self):
        upload = self.make_upload()

        result = create_raw_chargement_derived_import_reviews(
            upload,
            source_system_code="AIO_WEB",
            uploaded_by=self.user,
            period_start="2026-08-01",
            period_end="2026-08-17",
            original_filename="aio_mixed_raw.xlsx",
        )

        self.assertEqual(
            ImportSourceUpload.objects.count(),
            1,
        )

        self.assertEqual(
            ImportBatch.objects.count(),
            2,
        )

        source_upload = result.source_upload

        self.assertEqual(
            source_upload.source_system,
            self.source_system,
        )

        batches = {
            batch.brand.code: batch
            for batch in result.batches
        }

        self.assertEqual(
            set(batches),
            {"DELISKY", "NITA"},
        )

        delisky_batch = batches["DELISKY"]
        nita_batch = batches["NITA"]

        self.assertEqual(
            delisky_batch.source_upload_id,
            source_upload.pk,
        )
        self.assertEqual(
            nita_batch.source_upload_id,
            source_upload.pk,
        )

        self.assertEqual(
            delisky_batch.file_sha256,
            "",
        )
        self.assertEqual(
            nita_batch.file_sha256,
            "",
        )

        self.assertFalse(
            delisky_batch.source_file
        )
        self.assertFalse(
            nita_batch.source_file
        )

        self.assertEqual(
            delisky_batch.rows.count(),
            1,
        )
        self.assertEqual(
            nita_batch.rows.count(),
            1,
        )

        self.assertEqual(
            delisky_batch.rows.get().raw_data["VAN"],
            "DELISKY LIV01",
        )

        self.assertEqual(
            nita_batch.rows.get().raw_data["VAN"],
            "NITA LIV01",
        )


    def test_reprocessing_same_mixed_raw_file_reuses_derived_batches(self):
        first_upload = self.make_upload()

        file_bytes = first_upload.read()
        first_upload.seek(0)

        first = create_raw_chargement_derived_import_reviews(
            first_upload,
            source_system_code="AIO_WEB",
            uploaded_by=self.user,
            period_start="2026-08-01",
            period_end="2026-08-17",
            original_filename="aio_mixed_raw.xlsx",
        )

        first_batch_ids = {
            batch.brand.code: batch.pk
            for batch in first.batches
        }

        second_upload = SimpleUploadedFile(
            "aio_mixed_raw.xlsx",
            file_bytes,
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        second = create_raw_chargement_derived_import_reviews(
            second_upload,
            source_system_code="AIO_WEB",
            uploaded_by=self.user,
            period_start="2026-08-01",
            period_end="2026-08-17",
            original_filename="aio_mixed_raw.xlsx",
        )

        second_batch_ids = {
            batch.brand.code: batch.pk
            for batch in second.batches
        }

        self.assertEqual(
            ImportSourceUpload.objects.count(),
            1,
        )

        self.assertEqual(
            ImportBatch.objects.count(),
            2,
        )

        self.assertEqual(
            first.source_upload.pk,
            second.source_upload.pk,
        )

        self.assertEqual(
            first_batch_ids,
            second_batch_ids,
        )

        self.assertEqual(
            set(second_batch_ids),
            {"DELISKY", "NITA"},
        )

        self.assertEqual(
            ImportBatch.objects.get(
                brand=self.delisky
            ).rows.count(),
            1,
        )

        self.assertEqual(
            ImportBatch.objects.get(
                brand=self.nita
            ).rows.count(),
            1,
        )


    def test_reprocessing_does_not_modify_approved_derived_batch(self):
        from apps.imports.services.batch_review import (
            ImportBatchReviewError,
        )

        first_upload = self.make_upload()

        file_bytes = first_upload.read()
        first_upload.seek(0)

        first = create_raw_chargement_derived_import_reviews(
            first_upload,
            source_system_code="AIO_WEB",
            uploaded_by=self.user,
            period_start="2026-08-01",
            period_end="2026-08-17",
            original_filename="aio_mixed_raw.xlsx",
        )

        delisky_batch = next(
            batch
            for batch in first.batches
            if batch.brand.code == "DELISKY"
        )

        original_content_sha256 = (
            delisky_batch.content_sha256
        )

        delisky_batch.status = "APPROVED"
        delisky_batch.save(
            update_fields=["status"]
        )

        second_upload = SimpleUploadedFile(
            "aio_mixed_raw.xlsx",
            file_bytes,
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            ImportBatchReviewError
        ) as captured:
            create_raw_chargement_derived_import_reviews(
                second_upload,
                source_system_code="AIO_WEB",
                uploaded_by=self.user,
                period_start="2026-08-01",
                period_end="2026-08-17",
                original_filename="aio_mixed_raw.xlsx",
            )

        self.assertEqual(
            captured.exception.code,
            "immutable_batch",
        )

        delisky_batch.refresh_from_db()

        self.assertEqual(
            delisky_batch.status,
            "APPROVED",
        )

        self.assertEqual(
            delisky_batch.content_sha256,
            original_content_sha256,
        )

        self.assertEqual(
            ImportSourceUpload.objects.count(),
            1,
        )

        self.assertEqual(
            ImportBatch.objects.count(),
            2,
        )


    def test_reprocessing_rolls_back_mutable_batch_if_another_batch_is_immutable(self):
        from apps.imports.services.batch_review import (
            ImportBatchReviewError,
        )

        first_upload = self.make_upload()

        file_bytes = first_upload.read()
        first_upload.seek(0)

        first = create_raw_chargement_derived_import_reviews(
            first_upload,
            source_system_code="AIO_WEB",
            uploaded_by=self.user,
            period_start="2026-08-01",
            period_end="2026-08-17",
            original_filename="aio_mixed_raw.xlsx",
        )

        batches = {
            batch.brand.code: batch
            for batch in first.batches
        }

        delisky_batch = batches["DELISKY"]
        nita_batch = batches["NITA"]

        delisky_batch.review_summary = {
            "sentinel": "must-survive-rollback",
        }
        delisky_batch.save(
            update_fields=["review_summary"]
        )

        nita_batch.status = "APPROVED"
        nita_batch.save(
            update_fields=["status"]
        )

        second_upload = SimpleUploadedFile(
            "aio_mixed_raw.xlsx",
            file_bytes,
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        with self.assertRaises(
            ImportBatchReviewError
        ) as captured:
            create_raw_chargement_derived_import_reviews(
                second_upload,
                source_system_code="AIO_WEB",
                uploaded_by=self.user,
                period_start="2026-08-01",
                period_end="2026-08-17",
                original_filename="aio_mixed_raw.xlsx",
            )

        self.assertEqual(
            captured.exception.code,
            "immutable_batch",
        )

        delisky_batch.refresh_from_db()
        nita_batch.refresh_from_db()

        self.assertEqual(
            delisky_batch.review_summary,
            {
                "sentinel": "must-survive-rollback",
            },
        )

        self.assertEqual(
            nita_batch.status,
            "APPROVED",
        )

        self.assertEqual(
            ImportSourceUpload.objects.count(),
            1,
        )

        self.assertEqual(
            ImportBatch.objects.count(),
            2,
        )
