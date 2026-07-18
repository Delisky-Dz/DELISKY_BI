from __future__ import annotations

from dataclasses import dataclass
from os import PathLike
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class ReportRowReadError(Exception):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        details: dict[str, Any] | None = None,
    ):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


@dataclass(frozen=True, slots=True)
class RawReportRow:
    row_number: int
    values: tuple[tuple[str, Any], ...]

    def as_dict(self) -> dict[str, Any]:
        return dict(self.values)


@dataclass(frozen=True, slots=True)
class ReportRowReadResult:
    filename: str
    report_type: str
    worksheet_name: str
    headers: tuple[str, ...]
    rows: tuple[RawReportRow, ...]

    @property
    def row_count(self) -> int:
        return len(self.rows)


def _is_path_source(source: Any) -> bool:
    return isinstance(source, (str, PathLike))


def _remember_position(source: Any) -> int | None:
    if _is_path_source(source):
        return None

    if hasattr(source, "tell"):
        try:
            return int(source.tell())
        except (OSError, ValueError):
            return None

    return None


def _seek_start(source: Any) -> None:
    if not _is_path_source(source) and hasattr(source, "seek"):
        source.seek(0)


def _restore_position(
    source: Any,
    position: int | None,
) -> None:
    if (
        position is not None
        and not _is_path_source(source)
        and hasattr(source, "seek")
    ):
        try:
            source.seek(position)
        except (OSError, ValueError):
            pass


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True

    if isinstance(value, str):
        return not value.strip()

    return False


def _is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(
        _is_blank_value(value)
        for value in row
    )


def read_report_rows(
    source: Any,
    preflight_result: Any,
) -> ReportRowReadResult:
    if not preflight_result.is_valid:
        raise ReportRowReadError(
            "invalid_preflight",
            (
                "Rows cannot be read because the import "
                "preflight contains blocking errors."
            ),
        )

    parsed = preflight_result.parsed_filename
    inspection = preflight_result.inspection
    schema_validation = (
        preflight_result.schema_validation
    )

    if (
        parsed is None
        or inspection is None
        or schema_validation is None
    ):
        raise ReportRowReadError(
            "incomplete_preflight",
            (
                "The import preflight result is incomplete."
            ),
        )

    if inspection.worksheet_count != 1:
        raise ReportRowReadError(
            "unexpected_worksheet_count",
            "Exactly one worksheet is required.",
            details={
                "actual": inspection.worksheet_count,
            },
        )

    worksheet_inspection = inspection.worksheets[0]
    worksheet_name = (
        schema_validation.worksheet_name
        or worksheet_inspection.name
    )

    column_positions = (
        schema_validation.column_positions
    )

    if not column_positions:
        raise ReportRowReadError(
            "missing_column_positions",
            (
                "No validated column positions are "
                "available."
            ),
        )

    headers = tuple(
        header
        for header, _position in column_positions
    )

    original_position = _remember_position(source)
    workbook = None

    try:
        _seek_start(source)

        workbook = load_workbook(
            filename=source,
            read_only=True,
            data_only=True,
            keep_links=False,
        )

        try:
            worksheet = workbook[worksheet_name]
        except KeyError as exc:
            raise ReportRowReadError(
                "worksheet_not_found",
                (
                    "The validated worksheet no longer "
                    "exists in the workbook."
                ),
                details={
                    "worksheet": worksheet_name,
                },
            ) from exc

        raw_rows: list[RawReportRow] = []

        first_data_row = (
            worksheet_inspection.header_row_number + 1
        )

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=first_data_row,
                values_only=True,
            ),
            start=first_data_row,
        ):
            row_values = tuple(row)

            if _is_blank_row(row_values):
                continue

            mapped_values: list[
                tuple[str, Any]
            ] = []

            for header, position in column_positions:
                index = position - 1

                value = (
                    row_values[index]
                    if index < len(row_values)
                    else None
                )

                mapped_values.append(
                    (header, value)
                )

            raw_rows.append(
                RawReportRow(
                    row_number=row_number,
                    values=tuple(mapped_values),
                )
            )

        expected_row_count = (
            worksheet_inspection.data_row_count
        )

        if len(raw_rows) != expected_row_count:
            raise ReportRowReadError(
                "row_count_mismatch",
                (
                    "The workbook content changed after "
                    "the preflight inspection."
                ),
                details={
                    "expected": expected_row_count,
                    "actual": len(raw_rows),
                },
            )

        return ReportRowReadResult(
            filename=preflight_result.filename,
            report_type=parsed.report_type,
            worksheet_name=worksheet_name,
            headers=headers,
            rows=tuple(raw_rows),
        )

    except ReportRowReadError:
        raise
    except (
        InvalidFileException,
        KeyError,
        OSError,
        ValueError,
    ) as exc:
        raise ReportRowReadError(
            "workbook_read_failed",
            (
                "The workbook could not be read after "
                "preflight validation."
            ),
            details={
                "filename": preflight_result.filename,
            },
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()

        _restore_position(
            source,
            original_position,
        )
