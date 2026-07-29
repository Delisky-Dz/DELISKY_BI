from dataclasses import dataclass
from os import PathLike
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, is_zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


SUPPORTED_EXTENSIONS = {".xlsx"}
DEFAULT_MAX_FILE_SIZE_BYTES = 25 * 1024 * 1024


class ExcelInspectionError(Exception):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        details: dict[str, Any] | None = None,
    ):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


@dataclass(frozen=True, slots=True)
class WorksheetInspection:
    name: str
    header_row_number: int
    headers: tuple[str, ...]
    column_count: int
    data_row_count: int
    blank_row_count: int
    empty_header_positions: tuple[int, ...]
    duplicate_headers: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class WorkbookInspection:
    filename: str
    file_size_bytes: int
    worksheet_count: int
    worksheets: tuple[WorksheetInspection, ...]


def _is_path_source(source: Any) -> bool:
    return isinstance(source, (str, PathLike))


def _get_filename(
    source: Any,
    original_filename: str | None,
) -> str:
    if original_filename:
        return Path(original_filename).name

    if _is_path_source(source):
        return Path(source).name

    source_name = getattr(source, "name", "")
    if source_name:
        return Path(str(source_name)).name

    raise ExcelInspectionError(
        "missing_filename",
        "تعذر تحديد اسم ملف المصدر.",
    )


def _get_file_size(source: Any) -> int:
    if _is_path_source(source):
        return Path(source).stat().st_size

    declared_size = getattr(source, "size", None)
    if declared_size is not None:
        return int(declared_size)

    if not all(
        hasattr(source, attribute)
        for attribute in ("tell", "seek")
    ):
        raise ExcelInspectionError(
            "unreadable_source",
            "تعذر قراءة مصدر الملف المرفوع.",
        )

    current_position = source.tell()

    try:
        source.seek(0, 2)
        return int(source.tell())
    finally:
        source.seek(current_position)


def _remember_position(source: Any) -> int | None:
    if _is_path_source(source):
        return None

    if hasattr(source, "tell"):
        try:
            return int(source.tell())
        except (OSError, ValueError):
            return None

    return None


def _seek_start(source: Any) -> None:
    if not _is_path_source(source) and hasattr(source, "seek"):
        source.seek(0)


def _restore_position(
    source: Any,
    position: int | None,
) -> None:
    if (
        position is not None
        and not _is_path_source(source)
        and hasattr(source, "seek")
    ):
        try:
            source.seek(position)
        except (OSError, ValueError):
            pass


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True

    if isinstance(value, str):
        return not value.strip()

    return False


def _is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(_is_blank_value(value) for value in row)


def _header_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _trim_trailing_empty_headers(
    headers: list[str],
) -> list[str]:
    while headers and not headers[-1]:
        headers.pop()

    return headers


def _find_duplicate_headers(
    headers: list[str],
) -> tuple[str, ...]:
    seen: set[str] = set()
    duplicates: list[str] = []

    for header in headers:
        if not header:
            continue

        normalized = " ".join(header.split()).casefold()

        if normalized in seen and header not in duplicates:
            duplicates.append(header)
        else:
            seen.add(normalized)

    return tuple(duplicates)


def _inspect_worksheet(worksheet) -> WorksheetInspection:
    header_row_number = 0
    headers: list[str] = []
    data_row_count = 0
    blank_row_count = 0
    pending_blank_rows = 0

    for row_number, row in enumerate(
        worksheet.iter_rows(values_only=True),
        start=1,
    ):
        row_values = tuple(row)

        if header_row_number == 0:
            if _is_blank_row(row_values):
                continue

            header_row_number = row_number
            headers = _trim_trailing_empty_headers(
                [_header_text(value) for value in row_values]
            )
            continue

        if _is_blank_row(row_values):
            pending_blank_rows += 1
        else:
            blank_row_count += pending_blank_rows
            pending_blank_rows = 0
            data_row_count += 1

    if header_row_number == 0 or not headers:
        raise ExcelInspectionError(
            "missing_header",
            (
                f"ورقة Excel «{worksheet.title}» "
                "لا تحتوي على صف عناوين صالح.",
            ),
            details={
                "worksheet": worksheet.title,
            },
        )

    empty_header_positions = tuple(
        index
        for index, header in enumerate(headers, start=1)
        if not header
    )

    return WorksheetInspection(
        name=worksheet.title,
        header_row_number=header_row_number,
        headers=tuple(headers),
        column_count=len(headers),
        data_row_count=data_row_count,
        blank_row_count=blank_row_count,
        empty_header_positions=empty_header_positions,
        duplicate_headers=_find_duplicate_headers(headers),
    )


def inspect_excel_file(
    source: Any,
    *,
    original_filename: str | None = None,
    max_file_size_bytes: int = DEFAULT_MAX_FILE_SIZE_BYTES,
) -> WorkbookInspection:
    filename = _get_filename(source, original_filename)
    extension = Path(filename).suffix.lower()

    if extension not in SUPPORTED_EXTENSIONS:
        raise ExcelInspectionError(
            "unsupported_extension",
            "نوع الملف غير مدعوم. الامتداد المقبول هو .xlsx.",
            details={
                "filename": filename,
                "extension": extension,
            },
        )

    try:
        file_size_bytes = _get_file_size(source)
    except OSError as exc:
        raise ExcelInspectionError(
            "file_access_error",
            "تعذر الوصول إلى الملف المرفوع.",
        ) from exc

    if file_size_bytes <= 0:
        raise ExcelInspectionError(
            "empty_file",
            "الملف المرفوع فارغ.",
        )

    if file_size_bytes > max_file_size_bytes:
        raise ExcelInspectionError(
            "file_too_large",
            "حجم ملف Excel يتجاوز الحد المسموح.",
            details={
                "file_size_bytes": file_size_bytes,
                "max_file_size_bytes": max_file_size_bytes,
            },
        )

    original_position = _remember_position(source)
    workbook = None

    try:
        _seek_start(source)

        if not is_zipfile(source):
            raise ExcelInspectionError(
                "invalid_xlsx",
                "الملف ليس ملف Excel صالحًا بصيغة .xlsx.",
            )

        _seek_start(source)

        workbook = load_workbook(
            filename=source,
            read_only=True,
            data_only=True,
            keep_links=False,
        )

        worksheets = tuple(
            _inspect_worksheet(worksheet)
            for worksheet in workbook.worksheets
        )

        if not worksheets:
            raise ExcelInspectionError(
                "empty_workbook",
                "ملف Excel لا يحتوي على أي ورقة.",
            )

        return WorkbookInspection(
            filename=filename,
            file_size_bytes=file_size_bytes,
            worksheet_count=len(worksheets),
            worksheets=worksheets,
        )

    except ExcelInspectionError:
        raise
    except (
        BadZipFile,
        InvalidFileException,
        KeyError,
        OSError,
        ValueError,
    ) as exc:
        raise ExcelInspectionError(
            "invalid_xlsx",
            "تعذر فتح ملف Excel أو أن محتواه غير صالح.",
            details={
                "filename": filename,
            },
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()

        _restore_position(source, original_position)
