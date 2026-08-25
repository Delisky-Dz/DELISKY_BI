from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from apps.imports.models import (
    ImportSourceSystem,
    SourceProductPackaging,
)


REQUIRED_HEADERS = {
    "Num",
    "Désignation",
    "Colisage",
    "Activé",
}


class StockProductPackagingImportError(Exception):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        details: dict[str, Any] | None = None,
    ):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


@dataclass(frozen=True, slots=True)
class PackagingReviewItem:
    excel_row_number: int
    source_product_code: str
    designation: str
    reason: str


@dataclass(frozen=True, slots=True)
class StockProductPackagingImportResult:
    source_system_code: str
    total_rows: int
    created_count: int
    updated_count: int
    unchanged_count: int
    review_required_count: int
    review_items: tuple[PackagingReviewItem, ...]


def _normalize_text(
    value: Any,
) -> str:
    if value is None:
        return ""

    return " ".join(
        str(value)
        .replace("\xa0", " ")
        .split()
    )


def _normalize_product_code(
    value: Any,
) -> str:
    if value is None:
        return ""

    if isinstance(value, int):
        return str(value)

    if (
        isinstance(value, float)
        and value.is_integer()
    ):
        return str(int(value))

    return _normalize_text(value).upper()


def _parse_units_per_carton(
    value: Any,
) -> int | None:
    if value is None:
        return None

    text = _normalize_text(value)

    if not text:
        return None

    text = text.replace(",", ".")

    try:
        parsed = Decimal(text)
    except InvalidOperation:
        return None

    if parsed <= 0:
        return None

    if parsed != parsed.to_integral_value():
        return None

    return int(parsed)


def _parse_active(
    value: Any,
) -> bool:
    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(value)

    text = _normalize_text(value).casefold()

    if text in {
        "false",
        "0",
        "no",
        "non",
        "faux",
    }:
        return False

    return True


def _header_map(
    worksheet,
) -> dict[str, int]:
    header_row = next(
        worksheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        )
    )

    headers = {
        _normalize_text(value): index
        for index, value
        in enumerate(header_row)
        if _normalize_text(value)
    }

    missing = sorted(
        REQUIRED_HEADERS - set(headers)
    )

    if missing:
        raise StockProductPackagingImportError(
            "missing_required_headers",
            (
                "The STOCK reference file is missing "
                "required columns."
            ),
            details={
                "missing_headers": missing,
            },
        )

    return headers


def import_stock_product_packaging_file(
    source: Any,
    *,
    source_system_code: str,
) -> StockProductPackagingImportResult:
    try:
        source_system = (
            ImportSourceSystem.objects.get(
                code__iexact=(
                    source_system_code.strip()
                ),
                is_active=True,
            )
        )
    except ImportSourceSystem.DoesNotExist as exc:
        raise StockProductPackagingImportError(
            "unknown_source_system",
            (
                "The import source system does not "
                "exist or is inactive."
            ),
            details={
                "source_system_code":
                    source_system_code,
            },
        ) from exc

    workbook = load_workbook(
        source,
        read_only=True,
        data_only=True,
    )

    try:
        if len(workbook.worksheets) != 1:
            raise StockProductPackagingImportError(
                "unexpected_worksheet_count",
                (
                    "The STOCK reference file must "
                    "contain exactly one worksheet."
                ),
                details={
                    "worksheet_count":
                        len(workbook.worksheets),
                },
            )

        worksheet = workbook.worksheets[0]
        headers = _header_map(
            worksheet
        )

        rows_to_import = []
        seen_codes: set[str] = set()

        for excel_row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            if all(
                value is None
                for value in row
            ):
                continue

            def value_for(header: str):
                index = headers.get(header)

                if index is None:
                    return None

                if index >= len(row):
                    return None

                return row[index]

            source_product_code = (
                _normalize_product_code(
                    value_for("Num")
                )
            )

            designation = _normalize_text(
                value_for("Désignation")
            )

            if not source_product_code:
                raise StockProductPackagingImportError(
                    "missing_source_product_code",
                    (
                        "A STOCK product row has no "
                        "Num value."
                    ),
                    details={
                        "excel_row_number":
                            excel_row_number,
                    },
                )

            if not designation:
                raise StockProductPackagingImportError(
                    "missing_designation",
                    (
                        "A STOCK product row has no "
                        "Désignation value."
                    ),
                    details={
                        "excel_row_number":
                            excel_row_number,
                        "source_product_code":
                            source_product_code,
                    },
                )

            if source_product_code in seen_codes:
                raise StockProductPackagingImportError(
                    "duplicate_source_product_code",
                    (
                        "The same product Num appears "
                        "more than once in the STOCK file."
                    ),
                    details={
                        "excel_row_number":
                            excel_row_number,
                        "source_product_code":
                            source_product_code,
                    },
                )

            seen_codes.add(
                source_product_code
            )

            units_per_carton = (
                _parse_units_per_carton(
                    value_for("Colisage")
                )
            )

            rows_to_import.append(
                {
                    "excel_row_number":
                        excel_row_number,
                    "source_product_code":
                        source_product_code,
                    "barcode":
                        _normalize_text(
                            value_for("Barcode")
                        ).upper(),
                    "reference":
                        _normalize_text(
                            value_for("Réf")
                        ).upper(),
                    "designation":
                        designation,
                    "units_per_carton":
                        units_per_carton,
                    "needs_review":
                        units_per_carton is None,
                    "is_active":
                        _parse_active(
                            value_for("Activé")
                        ),
                }
            )

    finally:
        workbook.close()

    created_count = 0
    updated_count = 0
    unchanged_count = 0
    review_items = []

    with transaction.atomic():
        for item in rows_to_import:
            product = (
                SourceProductPackaging.objects
                .filter(
                    source_system=source_system,
                    source_product_code=(
                        item[
                            "source_product_code"
                        ]
                    ),
                )
                .first()
            )

            values = {
                "barcode":
                    item["barcode"],
                "reference":
                    item["reference"],
                "designation":
                    item["designation"],
                "units_per_carton":
                    item["units_per_carton"],
                "needs_review":
                    item["needs_review"],
                "is_active":
                    item["is_active"],
            }

            if product is None:
                product = SourceProductPackaging(
                    source_system=source_system,
                    source_product_code=(
                        item[
                            "source_product_code"
                        ]
                    ),
                    **values,
                )

                product.full_clean()
                product.save()

                created_count += 1

            else:
                changed = False

                for field_name, value in (
                    values.items()
                ):
                    if (
                        getattr(
                            product,
                            field_name,
                        )
                        != value
                    ):
                        setattr(
                            product,
                            field_name,
                            value,
                        )
                        changed = True

                if changed:
                    product.full_clean()
                    product.save()

                    updated_count += 1
                else:
                    unchanged_count += 1

            if item["needs_review"]:
                review_items.append(
                    PackagingReviewItem(
                        excel_row_number=(
                            item[
                                "excel_row_number"
                            ]
                        ),
                        source_product_code=(
                            item[
                                "source_product_code"
                            ]
                        ),
                        designation=(
                            item[
                                "designation"
                            ]
                        ),
                        reason=(
                            "unknown_units_per_carton"
                        ),
                    )
                )

    return StockProductPackagingImportResult(
        source_system_code=(
            source_system.code
        ),
        total_rows=len(
            rows_to_import
        ),
        created_count=created_count,
        updated_count=updated_count,
        unchanged_count=unchanged_count,
        review_required_count=len(
            review_items
        ),
        review_items=tuple(
            review_items
        ),
    )