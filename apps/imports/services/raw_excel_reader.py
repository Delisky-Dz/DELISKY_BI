from dataclasses import dataclass
from os import PathLike
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .excel_reader import (
    ExcelInspectionError,
    inspect_excel_file,
)


class RawExcelReadError(Exception):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        details: dict[str, Any] | None = None,
    ):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


@dataclass(frozen=True, slots=True)
class RawExcelRow:
    row_number: int
    values: tuple[tuple[str, Any], ...]

    def as_dict(self) -> dict[str, Any]:
        return dict(self.values)


@dataclass(frozen=True, slots=True)
class RawExcelReadResult:
    filename: str
    worksheet_name: str
    header_row_number: int
    headers: tuple[str, ...]
    rows: tuple[RawExcelRow, ...]


def _is_path_source(source: Any) -> bool:
    return isinstance(source, (str, PathLike))


def _remember_position(source: Any) -> int | None:
    if _is_path_source(source):
        return None

    if hasattr(source, "tell"):
        try:
            return int(source.tell())
        except (OSError, ValueError):
            return None

    return None


def _seek_start(source: Any) -> None:
    if not _is_path_source(source) and hasattr(source, "seek"):
        source.seek(0)


def _restore_position(
    source: Any,
    position: int | None,
) -> None:
    if (
        position is not None
        and not _is_path_source(source)
        and hasattr(source, "seek")
    ):
        try:
            source.seek(position)
        except (OSError, ValueError):
            pass


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True

    if isinstance(value, str):
        return not value.strip()

    return False


def _is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(
        _is_blank_value(value)
        for value in row
    )


def read_raw_excel_rows(
    source: Any,
    *,
    original_filename: str | None = None,
) -> RawExcelReadResult:
    try:
        inspection = inspect_excel_file(
            source,
            original_filename=original_filename,
        )
    except ExcelInspectionError as exc:
        raise RawExcelReadError(
            "excel_inspection_failed",
            "The raw Excel file failed inspection.",
            details={
                "cause_code": exc.code,
                "cause_details": dict(exc.details),
            },
        ) from exc

    if inspection.worksheet_count != 1:
        raise RawExcelReadError(
            "unexpected_worksheet_count",
            "Exactly one worksheet is required.",
            details={
                "actual": inspection.worksheet_count,
            },
        )

    worksheet_info = inspection.worksheets[0]

    if worksheet_info.duplicate_headers:
        raise RawExcelReadError(
            "duplicate_headers",
            (
                "The raw Excel worksheet contains "
                "duplicate column headers."
            ),
            details={
                "worksheet": worksheet_info.name,
                "duplicate_headers": list(
                    worksheet_info.duplicate_headers
                ),
            },
        )

    if worksheet_info.empty_header_positions:
        raise RawExcelReadError(
            "empty_headers",
            (
                "The raw Excel worksheet contains "
                "empty column headers."
            ),
            details={
                "worksheet": worksheet_info.name,
                "positions": list(
                    worksheet_info.empty_header_positions
                ),
            },
        )

    original_position = _remember_position(source)
    workbook = None

    try:
        _seek_start(source)

        workbook = load_workbook(
            filename=source,
            read_only=True,
            data_only=True,
            keep_links=False,
        )

        worksheet = workbook[
            worksheet_info.name
        ]

        headers = worksheet_info.headers
        rows: list[RawExcelRow] = []

        first_data_row = (
            worksheet_info.header_row_number + 1
        )

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=first_data_row,
                values_only=True,
            ),
            start=first_data_row,
        ):
            row_values = tuple(row)

            if _is_blank_row(row_values):
                continue

            values = tuple(
                (
                    header,
                    (
                        row_values[index]
                        if index < len(row_values)
                        else None
                    ),
                )
                for index, header in enumerate(headers)
            )

            rows.append(
                RawExcelRow(
                    row_number=row_number,
                    values=values,
                )
            )

        return RawExcelReadResult(
            filename=inspection.filename,
            worksheet_name=worksheet_info.name,
            header_row_number=(
                worksheet_info.header_row_number
            ),
            headers=headers,
            rows=tuple(rows),
        )

    except RawExcelReadError:
        raise
    except (
        InvalidFileException,
        KeyError,
        OSError,
        ValueError,
    ) as exc:
        raise RawExcelReadError(
            "workbook_read_failed",
            "The raw Excel workbook could not be read.",
            details={
                "filename": inspection.filename,
            },
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()

        _restore_position(
            source,
            original_position,
        )
